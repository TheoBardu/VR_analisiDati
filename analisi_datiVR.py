import pandas as pd
import openpyxl as ex
from openpyxl.styles import PatternFill
from numpy import zeros, arange, mean, std, max, round, ones, log10, sum, dot, sqrt
from os import chdir, getcwd, error, path
import re



# file = '/Users/theo/Desktop/Ermes/Misure/misF/misF/LSOURCES - Copia.txt'

        

class exel_file:
    '''
    Classe che serve per la manipolazione dei file excel
    '''

    def formatting_excel_VR8h_totale(file_path: str, output_path: str = None) -> str:
        """
        Formatta un file Excel applicando le seguenti modifiche su tutti i fogli:
        - Larghezza colonne adattata al contenuto
        - Colonna K (dalla riga 2) colorata in arancione chiaro (#FFA07A)
        - Colonna M (dalla riga 2) colorata in rosso scuro (#FF4500)
        - Colonna J (dalla riga 2) colorata in azzurro (#87CEEB)

        Args:
            file_path:   Percorso del file Excel di input.
            output_path: Percorso di output (opzionale).
                        Se None, sovrascrive il file originale.

        Returns:
            Percorso del file salvato.
        """
        import shutil
        import openpyxl
        from openpyxl.utils import get_column_letter
        from builtins import max as builtin_max

        
        if not path.isfile(file_path):
            raise FileNotFoundError(f"File non trovato: {file_path}")

        # Percorso di output
        if output_path is None:
            output_path = file_path
        elif output_path != file_path:
            shutil.copy2(file_path, output_path)

        wb = openpyxl.load_workbook(output_path)

        # Colori di riempimento
        fill_K = PatternFill(fill_type="solid", fgColor="FFA07A")  # arancione
        fill_M = PatternFill(fill_type="solid", fgColor="FF4500")  # rosso scuro
        fill_J = PatternFill(fill_type="solid", fgColor="87CEEB")  # azzurro

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # ── 1. DIMENSIONAMENTO AUTOMATICO COLONNE ─────────────────────────
            # Per ogni colonna calcola la lunghezza massima del contenuto
            col_widths: dict[int, float] = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # Lunghezza del valore come stringa
                        content_len = len(str(cell.value))
                        col_idx = cell.column          # intero 1-based
                        col_widths[col_idx] = builtin_max(col_widths.get(col_idx, 0), content_len)

            for col_idx, width in col_widths.items():
                col_letter = get_column_letter(col_idx)
                # Aggiungi un piccolo margine e imposta un minimo di 8
                ws.column_dimensions[col_letter].width = builtin_max(width + 2, 8)

            # ── 2. COLORI COLONNE (dalla riga 2 in poi) ───────────────────────
            color_map = {
                10: fill_J,   # colonna J
                11: fill_K,   # colonna K
                13: fill_M,   # colonna M
            }

            for col_idx, fill in color_map.items():
                col_letter = get_column_letter(col_idx)
                for row_idx in range(2, ws.max_row + 1):
                    ws[f"{col_letter}{row_idx}"].fill = fill

        wb.save(output_path)
        print(f"File salvato in: {output_path}")
        return output_path
    
    def color_column(file_path, column_letters, colors):
        '''
        Funzione che colora una colonna specifica di un file Excel.
        
        INPUT:
            file_path = <str>, percorso del file Excel
            column_letter = <list str>, lettera della colonna da colorare (es. 'A', 'B', ecc.)
            color = < list str>, colore in formato esadecimale (es. 'FFFF00' per il giallo)
        '''
        wb = ex.load_workbook(file_path) #carica il file exel
        ws = wb.active #seleziona il fogio attivo 


        
        idx_color = 0
        for letter in column_letters:
            fill = PatternFill(start_color=colors[idx_color], end_color=colors[idx_color], fill_type="solid")
            for cell in ws[letter]:
                cell.fill = fill
            idx_color += 1

        wb.save(file_path)
        print(f'Colonna/e {column_letters} colorata/e con successo')
    

    def adjust_column_lenght(file_path, column_letters):
        '''
        Funzione che adatta la lunghezza delle colonne in un file Excel.
        '''
        wb = ex.load_workbook(file_path) #cario il file excel
        ws = wb.active #seleziona il foglio attivo

        for letter in column_letters:
            max_lenght = 0 #ausiliaria, per salvare la lunghezza massima del nome della colonna
            for cell in ws[letter]:
                if len(str(cell.value)) > max_lenght:
                    max_lenght = len(str(cell.value))

            ws.column_dimensions[letter].width = max_lenght
            wb.save(file_path)


    def color_cell_VR8h(file_path, column_names = ['Peak', 'Leq_max','GrOm'], colors = ['e8643c','e8bd3c', '18AB49']):
        '''
        Funzione che colora la cella del file in maniera opportuna in base al rischio
        'e8643c','e8bd3c', '18AB49' : rosso, giallo, verde
        '''
        wb = ex.load_workbook(file_path) #carica il file exel
        ws = wb.active #seleziona il fogio attivo 
        
        #trovo l'indice delle colonne
        col = {}
        for cell in ws[1]:
            if cell.value in column_names:
                col[cell.value] = cell.column
        



        # itero sul numero di righe per colorare le celle
        for i in range(2 , ws.max_row + 1):
        
            #rischio basso
            if ws.cell(row=i, column = col[column_names[1]]).value <= 80.0 and ws.cell(row=i, column = col[column_names[0]]).value <= 135.0:
                #creo il settaggio del riempimento verde
                fill = PatternFill(start_color=colors[2], end_color=colors[2], fill_type="solid")
                ws.cell(row=i,column= col[column_names[2]]).fill = fill
            
            #rischio medio
            if ws.cell(row=i, column = col[column_names[1]]).value > 80.0 and ws.cell(row=i, column = col[column_names[1]]).value <= 85.0 and ws.cell(row=i, column = col[column_names[0]]).value > 135.0 and ws.cell(row=i, column = col[column_names[0]]).value <= 137.0:
                #creo il settaggio del riempimento giallo
                fill = PatternFill(start_color=colors[1], end_color=colors[1], fill_type="solid")
                ws.cell(row=i,column= col[column_names[2]]).fill = fill
            
            #rischio alto
            if ws.cell(row=i, column = col[column_names[1]]).value > 85.0 and ws.cell(row=i, column = col[column_names[1]]).value <= 87.0 and ws.cell(row=i, column = col[column_names[0]]).value > 137.0 and ws.cell(row=i, column = col[column_names[0]]).value <= 140.0:
                #creo il settaggio del riempimento giallo
                fill = PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
                ws.cell(row=i,column= col[column_names[2]]).fill = fill
            
            #rischio fuori scala (Warning)
            if ws.cell(row=i, column = col[column_names[1]]).value > 87 or ws.cell(row=i, column = col[column_names[0]]).value > 140.0:
                fill = PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
                ws.cell(row=i,column= col[column_names[2]]).fill = fill
                print('########## Attenzione ##########\n####################\nPossibile errore nel calcolo oppure valore rumore troppo alto!\nControlla i dati.\n####################')


        
        
        
        wb.save(file_path)

    def colora_classe_rischio(path: str) -> None:
        from openpyxl import load_workbook
        COLORI = {
            "BASSA": "32CD32",
            "MEDIA": "00BFFF",
            "ALTA":  "B22222",
        }

        wb = load_workbook(path)
    
        for ws in wb.worksheets:
            col_idx = None
    
            # Cerca la colonna con intestazione "classe_rischio" nella riga 1
            for cell in ws[1]:
                if cell.value and str(cell.value).strip().lower() == "classe_rischio":
                    col_idx = cell.column
                    break
    
            if col_idx is None:
                print(f"Foglio '{ws.title}': colonna 'classe_rischio' non trovata, skip.")
                continue
    
            # Colora le celle dalla riga 2 in poi
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                valore = str(cell.value).strip().upper() if cell.value is not None else ""
                if valore in COLORI:
                    cell.fill = PatternFill(
                        fill_type="solid",
                        start_color=COLORI[valore],
                        end_color=COLORI[valore],
                    )
    
        wb.save(path)
        print(f"File salvato: {path}")

    def inserisci_valutazione_schede( path_riepilogo: str, path_totale: str, path_output: str,):
        """
        inserisci_valutazione_rumore.py

        Funzione che legge i dati dal file VR8h_riepilogo.xlsx e li inserisce
        nelle schede del file VR8h_totale.xlsx, aggiungendo la tabellina di
        valutazione a destra (colonne M:T, righe 2-6) di ogni foglio "Scheda N".

        Layout della tabellina inserita:
        Riga 2 (M2:T2 merged): "VALUTAZIONE SU BASE GIORNALIERA"  (header)
        Riga 3: LEX,8h | N3=Lex8h | ± | P3=U | → |   | LEX MAX = | T3=Lex_max
        Riga 4:                                             Massimo dei Lpicco,C = | T4=L_picco_C
        Righe 5-6 (M5:T6 merged): "CLASSE RISCHIO <classe_rischio>"  (colorato)
        """

        import copy
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.styles.colors import Color
        from openpyxl.utils import get_column_letter


        # ---------------------------------------------------------------------------
        # Palette colori classe_rischio (rgb in formato AARRGGBB / RRGGBB)
        # ---------------------------------------------------------------------------
        CLASSE_FILL = {
            "BASSA":  {"bg": "FF32CD32", "font_color": "FF000000"},  # verde lime
            "MEDIA":  {"bg": "FF00BFFF", "font_color": "FF000000"},  # azzurro
            "ALTA":   {"bg": "FFB22222", "font_color": "FFFFFFFF"},  # rosso scuro
        }

        # ---------------------------------------------------------------------------
        # Stili riutilizzabili (costruiti una sola volta)
        # ---------------------------------------------------------------------------
        BORDER_MEDIUM = Side(border_style="medium", color="FF000000")
        BORDER_NONE   = Side(border_style=None)

        def _border(**kwargs):
            sides = {k: kwargs.get(k, BORDER_NONE) for k in ("left","right","top","bottom")}
            return Border(**sides)

        def _font(bold=False, size=11, color=None, name="Calibri"):
            kw = dict(bold=bold, size=size, name=name)
            if color:
                kw["color"] = color
            return Font(**kw)

        def _fill(rgb):
            return PatternFill("solid", fgColor=Color(rgb=rgb))

        def _align(h="general", v="bottom", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


        # ---------------------------------------------------------------------------
        # Helpers
        # ---------------------------------------------------------------------------
        def _set(ws, row, col, value=None, *, font=None, fill=None, alignment=None, border=None):
            cell = ws.cell(row, col)
            if value is not None:
                cell.value = value
            if font      is not None: cell.font      = font
            if fill      is not None: cell.fill      = fill
            if alignment is not None: cell.alignment = alignment
            if border    is not None: cell.border    = border
            return cell


        def _safe_merge(ws, min_row, min_col, max_row, max_col):
            """Esegue merge solo se non già presente."""
            key = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            existing = {str(r) for r in ws.merged_cells.ranges}
            if key not in existing:
                ws.merge_cells(start_row=min_row, start_column=min_col,
                            end_row=max_row,   end_column=max_col)

        def _last_populated_col(ws) -> int:
            """Restituisce l'indice (1-based) dell'ultima colonna con almeno un valore."""
            max_col = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and cell.column > max_col:
                        max_col = cell.column
            return max_col if max_col > 0 else 1

        # ---------------------------------------------------------------------------
        # Funzione principale
        # ---------------------------------------------------------------------------
        def inserisci_valutazione(
            path_riepilogo: str,
            path_totale: str,
            path_output: str,
        ):
            """
            Legge i dati da `path_riepilogo` e li inserisce nelle schede di
            `path_totale`, salvando il risultato in `path_output`.

            Parametri
            ----------
            path_riepilogo : str
                Percorso di VR8h_riepilogo.xlsx
            path_totale : str
                Percorso di VR8h_totale.xlsx (file sorgente, non viene modificato)
            path_output : str
                Percorso del file Excel di output
            """

            # ------------------------------------------------------------------
            # 1. Leggi il riepilogo
            # ------------------------------------------------------------------
            wb_riepilogo = openpyxl.load_workbook(path_riepilogo, data_only=True)
            sh_riepilogo = wb_riepilogo.active

            # Costruiamo un dict: ID_GrOm -> {Lex8h, U, Lex_max, L_picco_C, classe_rischio}
            headers = [cell.value for cell in sh_riepilogo[1]]
            col_idx = {h: i for i, h in enumerate(headers)}

            riepilogo = {}
            for row in sh_riepilogo.iter_rows(min_row=2, values_only=True):
                id_grom = str(row[col_idx["ID_GrOm"]])
                riepilogo[id_grom] = {
                    "Lex8h":         row[col_idx["Lex8h"]],
                    "U":             row[col_idx["U"]],
                    "Lex_max":       row[col_idx["Lex_max"]],
                    "L_picco_C":     row[col_idx["L_picco_C"]],
                    "classe_rischio": row[col_idx["classe_rischio"]],
                }

            # ------------------------------------------------------------------
            # 2. Apri VR8h_totale e lavora su ogni scheda
            # ------------------------------------------------------------------
            wb = openpyxl.load_workbook(path_totale)

            for shname in wb.sheetnames:
                if not shname.startswith("Scheda"):
                    continue

                ws = wb[shname]

                # Recupera l'ID dal primo dato (riga 2, colonna A)
                id_grom = str(ws.cell(2, 1).value)
                if id_grom not in riepilogo:
                    print(f"  [AVVISO] '{shname}': ID '{id_grom}' non trovato nel riepilogo. Skip.")
                    continue

                dati = riepilogo[id_grom]
                lex8h        = dati["Lex8h"]
                u            = dati["U"]
                lex_max      = dati["Lex_max"]
                l_picco_c    = dati["L_picco_C"]
                classe       = (dati["classe_rischio"] or "").strip().upper()

                # Palette colore per la classe
                palette = CLASSE_FILL.get(classe, {"bg": "FFD3D3D3", "font_color": "FF000000"})
                fill_classe = _fill(palette["bg"])
                font_classe_color = palette["font_color"]

                # ------------------------------------------------------------------
                # Calcola colonna di partenza: ultima col popolata + 3
                # ------------------------------------------------------------------
                # MODIFICA: le 8 colonne della tabellina non sono più hardcodate a
                # M..T (13..20) ma calcolate dinamicamente in base al contenuto del
                # foglio. col_start = max_col_popolata + 3; col_end = col_start + 7.
                last_col  = _last_populated_col(ws)
                col_start = last_col + 3          # prima colonna della tabellina
                M = col_start
                N = col_start + 1   # Lex8h
                O = col_start + 2   # ±
                P = col_start + 3   # U
                Q = col_start + 4   # →
                R = col_start + 5   # (vuota)
                S = col_start + 6   # "LEX MAX =" / "Massimo dei Lpicco,C..."
                T = col_start + 7   # valore numerico (col_end)

                # ------------------------------------------------------------------
                # Riga 2: header "VALUTAZIONE SU BASE GIORNALIERA" (merge col_start:col_end)
                # ------------------------------------------------------------------
                # MODIFICA: la pulizia dei merge esistenti usa col_start/T calcolati
                for mr in list(ws.merged_cells.ranges):
                    if mr.min_col >= M and mr.max_col <= T and 2 <= mr.min_row <= 6:
                        ws.merged_cells.remove(mr)

                _safe_merge(ws, 2, M, 2, T)
                _set(ws, 2, M,
                    value="VALUTAZIONE SU BASE GIORNALIERA",
                    font=_font(bold=True, size=14),
                    fill=_fill("FFCCFFCC"),          # indexed 42 -> light green
                    alignment=_align("center"),
                    border=_border(left=BORDER_MEDIUM, right=BORDER_MEDIUM,
                                    top=BORDER_MEDIUM, bottom=BORDER_MEDIUM))
                # Celle N2:T2 (parte del merge) — solo bordi top/bottom per il contorno esterno
                for c in range(N, T + 1):
                    _set(ws, 2, c,
                        border=_border(top=BORDER_MEDIUM, bottom=BORDER_MEDIUM,
                                        right=BORDER_MEDIUM if c == T else BORDER_NONE))

                # ------------------------------------------------------------------
                # Riga 3: LEX,8h | Lex8h | ± | U | → |   | LEX MAX = | Lex_max
                # ------------------------------------------------------------------
                fill_bianco = _fill("FFFFFFFF")   # sfondo bianco (indexed 9)

                _set(ws, 3, M,
                    value="LEX,8h",
                    font=_font(bold=True, size=14),
                    fill=fill_bianco,
                    alignment=_align("center"),
                    border=_border(left=BORDER_MEDIUM))

                _set(ws, 3, N,
                    value=lex8h,
                    font=_font(bold=True, size=14),
                    fill=fill_bianco,
                    alignment=_align("right"))

                _set(ws, 3, O,
                    value="±",
                    font=_font(bold=True, size=14),
                    fill=fill_bianco,
                    alignment=_align("center"))

                _set(ws, 3, P,
                    value=u,
                    font=_font(bold=True, size=14),
                    fill=fill_bianco,
                    alignment=_align("left"))

                _set(ws, 3, Q,
                    value="→",
                    font=_font(bold=False, size=22),
                    fill=fill_bianco,
                    alignment=_align("right"))

                _set(ws, 3, R,
                    value=None,
                    font=_font(size=22),
                    fill=fill_bianco,
                    alignment=_align("left"))

                _set(ws, 3, S,
                    value="LEX MAX =",
                    font=_font(bold=True, size=14),
                    fill=fill_bianco,
                    alignment=_align("right"))

                _set(ws, 3, T,
                    value=lex_max,
                    font=_font(bold=True, size=14),
                    fill=fill_bianco,
                    alignment=_align("center"),
                    border=_border(right=BORDER_MEDIUM))

                # ------------------------------------------------------------------
                # Riga 4: (vuota a sinistra) | Massimo dei Lpicco,C misurati = | L_picco_C
                # ------------------------------------------------------------------
                for c in range(M, S):
                    _set(ws, 4, c,
                        value=None,
                        fill=fill_bianco,
                        font=_font(size=14),
                        border=_border(left=BORDER_MEDIUM if c == M else BORDER_NONE))

                _set(ws, 4, S,
                    value="Massimo dei Lpicco,C misurati =",
                    font=_font(bold=True, size=14),
                    fill=fill_bianco,
                    alignment=_align("right"))

                _set(ws, 4, T,
                    value=l_picco_c,
                    font=_font(bold=True, size=14),
                    fill=fill_bianco,
                    alignment=_align("center"),
                    border=_border(right=BORDER_MEDIUM))

                # ------------------------------------------------------------------
                # Righe 5-6: CLASSE RISCHIO (M5:T6 merged, colorato)
                # ------------------------------------------------------------------
                _safe_merge(ws, 5, M, 6, T)
                _set(ws, 5, M,
                    value=f"CLASSE RISCHIO {classe}",
                    font=_font(bold=True, size=14, color=font_classe_color),
                    fill=fill_classe,
                    alignment=_align("center", "center"),
                    border=_border(left=BORDER_MEDIUM, right=BORDER_MEDIUM,
                                    top=BORDER_MEDIUM, bottom=BORDER_MEDIUM))

                # Celle della seconda riga del merge (row 6) — bordi del contorno
                for c in range(N, T + 1):
                    _set(ws, 6, c,
                        border=_border(bottom=BORDER_MEDIUM,
                                        right=BORDER_MEDIUM if c == T else BORDER_NONE))

                # ------------------------------------------------------------------
                # Larghezze colonne M:T (se non già impostate)
                # ------------------------------------------------------------------
                col_widths = {M: 6, N: 7, O: 4, P: 6, Q: 4, R: 4, S: 28, T: 10}
                for col_num, width in col_widths.items():
                    col_letter = get_column_letter(col_num)
                    cd = ws.column_dimensions[col_letter]
                    if cd.width is None or cd.width < 1:
                        cd.width = width

                # Altezze righe (solo se non impostate)
                row_heights = {2: 19, 3: 28, 4: 19, 5: 16, 6: 16}
                for r, h in row_heights.items():
                    rd = ws.row_dimensions[r]
                    if rd.height is None or rd.height < 1:
                        rd.height = h

                print(f"  [OK] '{shname}': col_start={get_column_letter(M)} (last_col={get_column_letter(last_col)}+3), "
                    f"classe={classe}, Lex8h={lex8h}, U={u}, Lex_max={lex_max}, L_picco_C={l_picco_c}")

            # ------------------------------------------------------------------
            # 3. Salva
            # ------------------------------------------------------------------
            wb.save(path_output)
            print(f"\nFile salvato in: {path_output}")
        
        inserisci_valutazione(path_riepilogo, path_totale, path_output)


    def transfer_riepilogo2aggiornato(path_riepilogo: str, path_aggiornato: str) -> None:
        """
        Copia il foglio 'Riepilogo' di VR8h_riepilogo.xlsx come PRIMO foglio
        di VR8h_totale_aggiornato.xlsx, mantenendo valori e formattazione
        (in particolare i colori di classe_rischio impostati da colora_classe_rischio).

        Se il foglio 'Riepilogo' esiste già nel file aggiornato viene rimosso e
        riscritto, così la funzione è ri-eseguibile senza creare duplicati.

        Parametri
        ----------
        path_riepilogo : str
            Percorso di VR8h_riepilogo.xlsx
        path_aggiornato : str
            Percorso di VR8h_totale_aggiornato.xlsx (modificato in place)
        """
        import copy

        NOME_FOGLIO = 'Riepilogo'

        wb_r = ex.load_workbook(path_riepilogo)
        ws_r = wb_r[NOME_FOGLIO] if NOME_FOGLIO in wb_r.sheetnames else wb_r.worksheets[0]

        wb_a = ex.load_workbook(path_aggiornato)

        # Idempotenza: rimuovo un eventuale Riepilogo già presente
        if NOME_FOGLIO in wb_a.sheetnames:
            del wb_a[NOME_FOGLIO]

        ws_a = wb_a.create_sheet(NOME_FOGLIO, 0)  # indice 0 -> primo foglio

        # Celle: valori + stili
        for row in ws_r.iter_rows():
            for cell in row:
                new_cell = ws_a.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font          = copy.copy(cell.font)
                    new_cell.fill          = copy.copy(cell.fill)
                    new_cell.border        = copy.copy(cell.border)
                    new_cell.alignment     = copy.copy(cell.alignment)
                    new_cell.number_format = cell.number_format

        # Larghezze colonne e altezze righe
        for key, dim in ws_r.column_dimensions.items():
            ws_a.column_dimensions[key].width = dim.width
        for key, dim in ws_r.row_dimensions.items():
            ws_a.row_dimensions[key].height = dim.height

        # Celle unite
        for mr in ws_r.merged_cells.ranges:
            ws_a.merge_cells(str(mr))

        wb_a.save(path_aggiornato)
        print(f"Foglio '{NOME_FOGLIO}' inserito come primo foglio di: {path_aggiornato}")


    def colora_bordi_celle(path: str) -> None:
        """
        Applica bordi sottili neri alle celle popolate di tutti i fogli del workbook.

        - Foglio 'Riepilogo': borda l'intero rettangolo di celle popolate (A1 -> ultima
          riga/colonna con contenuto).
        - Fogli 'Scheda N': borda in maniera dinamica due aree distinte
            1. la tabella dati a sinistra (colonna A fino all'ultima colonna
               dell'intestazione, riga 1 fino all'ultima riga popolata);
            2. il blocco 'Analisi DPI in dotazione' (se presente), dal titolo fino
               all'ultima riga/colonna popolata del blocco.
          Il riquadro 'VALUTAZIONE SU BASE GIORNALIERA' non rientra in queste aree e
          conserva i suoi bordi 'medium'.

        La funzione è idempotente e non sovrascrive bordi di stile 'medium' già presenti.

        Parametri
        ----------
        path : str
            Percorso del file excel da bordare (modificato in place)
        """
        from openpyxl.styles import Border, Side
        from config import TESTO_TITOLO_DPI

        BORDER_THIN = Side(border_style='thin', color='FF000000')

        def _lato(lato_esistente, applica: bool):
            """Restituisce il lato da usare: conserva i bordi 'medium' preesistenti."""
            if lato_esistente is not None and lato_esistente.style == 'medium':
                return lato_esistente
            return BORDER_THIN if applica else lato_esistente

        def _applica_bordi(ws, min_row, min_col, max_row, max_col):
            """
            Borda il rettangolo indicato. Sulle celle unite disegna solo il perimetro
            esterno del range, lasciando vuoti i lati interni.
            """
            if max_row < min_row or max_col < min_col:
                return

            # Mappa cella -> range unito di appartenenza
            merges = list(ws.merged_cells.ranges)

            def _merge_di(r, c):
                for mr in merges:
                    if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
                        return mr
                return None

            for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                    min_col=min_col, max_col=max_col):
                for cell in row:
                    r, c = cell.row, cell.column
                    mr = _merge_di(r, c)
                    if mr is None:
                        left = top = right = bottom = True
                    else:
                        # solo i lati sul perimetro del merge
                        left   = (c == mr.min_col)
                        right  = (c == mr.max_col)
                        top    = (r == mr.min_row)
                        bottom = (r == mr.max_row)

                    b = cell.border
                    cell.border = Border(
                        left=_lato(b.left,   left),
                        right=_lato(b.right,  right),
                        top=_lato(b.top,    top),
                        bottom=_lato(b.bottom, bottom),
                    )

        def _ultima_riga_popolata(ws, min_col, max_col, from_row=1):
            """Ultima riga con almeno un valore nell'intervallo di colonne indicato."""
            last = 0
            for row in ws.iter_rows(min_row=from_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    if cell.value is not None:
                        last = cell.row
                        break
            return last

        def _ultima_col_popolata(ws, min_row, max_row, from_col=1):
            """Ultima colonna con almeno un valore nell'intervallo di righe indicato."""
            last = 0
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=from_col):
                for cell in row:
                    if cell.value is not None and cell.column > last:
                        last = cell.column
            return last

        def _ultima_col_contigua(ws, riga=1):
            """
            Ultima colonna dell'intestazione prima della prima cella vuota:
            serve a fermarsi al bordo della tabella dati senza includere i blocchi
            scritti più a destra (valutazione, DPI).
            """
            col = 0
            for cell in ws[riga]:
                if cell.value is None:
                    break
                col = cell.column
            return col

        wb = ex.load_workbook(path)

        for ws in wb.worksheets:
            if ws.max_row == 1 and ws.max_column == 1 and ws['A1'].value is None:
                continue  # foglio vuoto

            if ws.title.strip().lower().startswith('scheda'):
                # --- 1. Tabella dati a sinistra -------------------------------
                max_col_tab = _ultima_col_contigua(ws, riga=1)
                if max_col_tab > 0:
                    max_row_tab = _ultima_riga_popolata(ws, 1, max_col_tab)
                    _applica_bordi(ws, 1, 1, max_row_tab, max_col_tab)

                # --- 2. Blocco 'Analisi DPI in dotazione' ---------------------
                riga_dpi = col_dpi = None
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None and TESTO_TITOLO_DPI in str(cell.value):
                            riga_dpi, col_dpi = cell.row, cell.column
                            break
                    if riga_dpi is not None:
                        break

                if riga_dpi is not None:
                    max_row_dpi = _ultima_riga_popolata(ws, col_dpi, ws.max_column, from_row=riga_dpi)
                    max_col_dpi = _ultima_col_popolata(ws, riga_dpi, max_row_dpi, from_col=col_dpi)
                    _applica_bordi(ws, riga_dpi, col_dpi, max_row_dpi, max_col_dpi)

            else:
                # Riepilogo (e ogni altro foglio): tutto il blocco popolato
                max_col = _ultima_col_popolata(ws, 1, ws.max_row)
                if max_col > 0:
                    max_row = _ultima_riga_popolata(ws, 1, max_col)
                    _applica_bordi(ws, 1, 1, max_row, max_col)

        wb.save(path)
        print(f"Bordi applicati nel file: {path}")



class files:
    '''
    Classe che serve per leggere e scrivere i file
    '''

    def read_measure_file(file, letter_ID, format='csv',ntrack = 6, decimals = 1, read_version = '1'):
        '''
        Funzione che legge il file di misura in txt o csv e restituisce un dataframe pandas.

        INPUT:
            file = <str>, directory del file da leggere
            letter_ID = <str>, lettera della cartella dati (tipicamente D,E,F o W)
            format = <str>, "txt" o "csv" a seconda se leggere in txt o csv
            ntrack = <int>, default = 3. Solo per format = 'csv'. Numero di tracce di divisione del file intero.
            decimals = <int>, number of decimals in the data
            version = <str>, versione della modalità del file in lettura. 1 è la vecchia (legge da misE,..) 2 è la nuova (legge dalle sottocartelle di misE,..)
        OUTPUT:
            df = <pd.DataFrame>, dataframe contenente i dati della misura
        '''
        if format == 'txt':
            df = pd.DataFrame(columns=['fileID','letter_ID','nTrack','LeqA_min','LeqA_max','LeqA_eq','LeqC_min','LeqC_max','LeqC_eq','PeakC_max','PeakC_eq','durata','inizio','fine'])
            nFiles = 0 #numero di file nella misura
            

            #apertura e lettura del file
            with open(file, 'r', encoding='utf-16') as f:
                lines = f.readlines() #lettura delle linee

                
            #iterazione per leggere le singola linee
            for line in lines:
                    if 'File' in line:
                        nFiles += 1 #tengo conto del numero di misure presenti nel file
            
            # inizializzazione vettori
            # fileIDs = zeros(nFiles)
            # inizio = zeros(nFiles)
            # fine = zeros(nFiles)
            # nSorgenti = zeros(nFiles)
            fileIDs = []
            letter_IDs = []
            inizio=[]
            fine = []
            nSorgenti = []
            
            LeqA_min = [] # livello equalizzato A minimo 
            LeqA_max = [] # livello equalizzato A massimo
            LeqA_eq = [] # livello equalizzato A equivalente
            
            LeqC_min = [] # livello equalizzato C minimo
            LeqC_max = [] # livello equalizzato C massimo
            LeqC_eq = [] # livello equalizzato C equivalente

            PeakC_max = [] # picco C minimo
            PeakC_eq = [] # picco C massimo

            
            durata = [] # durata della traccia
            

            inxd_letter_IDs = 0 # tengo conto del numero di misura (es: D1, D2, D3, ...)
            indx_inizio = 0
            indx_fine = 0
            indx = 0
            for line in lines:
                #trovo dove sta la voce File
                if 'File' in line:
                    l0 = line.split() #questa è la riga del nome del file
                    # print(l)
                    # print(lines[indx])
                    # fileIDs[inxd_fileIDs] = l[1]
                    inxd_letter_IDs += 1
                    
                    l1 = lines[indx+1].split() # questa è la riga dell'inizio
                    

                    l2 = lines[indx+2].split() #questa è la riga della fine
                    
                    l3 = lines[indx+3].split() #questa è la riga del numero di sorgenti
                    
                    ntracks = arange(len(l3)-1) #creo un array con il numero di sorgenti in ordine crescente

                    #itero sul numero di sorgenti
                    for i in range(len(ntracks)):
                        fileIDs.append(l0[1]) #creo tante copie del nome sorgente quante sono le sorgenti
                        letter_IDs.append(letter_ID+str(inxd_letter_IDs)) #creo tante copie del letter ID  quante sono le sorgenti
                        inizio.append(l1[2]) # creo tante copie dell'inizio quante sono le sorgenti
                        fine.append(l2[2]) # creo tante copie della fine quante sono le sorgenti
                        nSorgenti.append(ntracks[i]+1) # creo un numero di sorgenti in ordine crescente

                    

                    #salvataggio livelli euqalizzati A e durata tracciati
                    l7 = lines[indx+7].split() #questa è la riga dei livelli LeqA
                    # print(l7)
                    for n in range(len(ntracks)):
                        LeqA_max.append(float(l7[5 + (n * 4)])); durata.append(l7[8 + (n*4)])
                        LeqA_min.append(float(l7[6 + (n * 4)]))
                        LeqA_eq.append(float(l7[7 + (n * 4)]))
            
                    # salvataggio livelli equalizzati C
                    l8 = lines[indx+8].split() #questa è la riga dei livelli C
                    # print(l8)
                    for n in range(len(ntracks)):
                        LeqC_max.append(float(l8[5 + (n * 4)]))
                        LeqC_min.append(float(l8[6 + (n * 4)]))
                        LeqC_eq.append(float(l8[7 + (n * 4)]))

                    #salvataggio picchi C
                    l9 = lines[indx+9].split()
                    # print(l9)
                    for n in range(len(ntracks)):
                        PeakC_eq.append(float(l9[5 + (n * 3)]))
                        PeakC_max.append(float(l9[6 + (n * 3)]))
                
                indx += 1

                

            df['fileID'] = fileIDs
            df['letter_ID'] = letter_IDs
            df['inizio'] = inizio
            df['fine'] = fine
            df['nTrack'] = nSorgenti
            df['LeqA_min'] = LeqA_min
            df['LeqA_max'] = LeqA_max
            df['LeqA_eq'] = LeqA_eq 
            df['LeqC_min'] = LeqC_min
            df['LeqC_max'] = LeqC_max
            df['LeqC_eq'] = LeqC_eq
            df['PeakC_max'] = PeakC_max
            df['PeakC_eq'] = PeakC_eq
            df['durata'] = durata
            # df['inizio'] = inizio
            # print(nFiles) #for debug
            print('Lettura e creazione dataFrame completata')
            return df, ntracks[-1], nFiles
        
        elif format == 'csv':
            import glob 
            from numpy import average, min, max
            
            
            print('Reading csv files only')

            #inizializzo il dataFrame pandas con il riassunto delle misure
            df_tot = pd.DataFrame(columns=['fileID','letter_ID','nTrack','LeqA_min','LeqA_max','LeqA_eq','LeqC_min','LeqC_max','LeqC_eq','PeakC_max','PeakC_eq','LASeq_T','LAIeq_T','durata','inizio','fine'])
            
            
            #Inizializzazione delle variabili di df_tot
            fileIDs = []
            letter_IDs = []
            inizio = []
            fine = []
            durata = []
            ntrack_id = []
            LeqA_min = []
            LeqA_max = []
            LeqA_eq = []
            LeqC_min = []
            LeqC_max = []
            LeqC_eq = []
            PeakC_max = []
            PeakC_eq = []
            LASeq_T = []
            LAIeq_T = []
            
            if read_version == '1':
                csv_files = glob.glob('*.csv') # salvo la lista di tutti i file csv che ci sono
    
            
            elif read_version == '2':
                # Prima di procedere alla lettura dei csv, cerco dove sono i file nelle sottocartelle
                all_subdirs = [
                    d for d in glob.glob('*')
                    if path.isdir(d) and re.search(r'_(\d+)$', d)
                ]
                # ordino i file in ordine crescente 0001, 0002, 0003, ...
                all_subdirs.sort(key=lambda d: int(re.search(r'_(\d+)$', d).group(1)))
                
                # Per ogni sottocartella, raccoglie i CSV al suo interno (ordinati per nome)
                csv_files = []
                for subdir in all_subdirs:
                    found = sorted(glob.glob(path.join(subdir, '*.csv')))
                    csv_files.extend(found)
                
            else:
                print("Errore: read_version non valida. Usa '1' (vecchia modalità) o '2' (nuova modalità).")
         

            # Procedo alla creazione del dataframe ============
            csv_files.sort() # riordino per nome la lista dei file

            letter_id_number = 1 #inizializzo il numero della misura ad 1 per ogni nuovo file misure ( in modo che d1,d2,d3...f1,f2,f3...)
            #itero sulla lista del numero di files csv
            for file in csv_files:
                
                df = pd.read_csv(file,encoding='latin', skiprows=1, sep=';', engine = 'python') #leggo il file csv delle misure
                df.iloc[:,0] = pd.to_datetime(df.iloc[:,0],format='%H:%M:%S', errors='coerce') # converto i dati della colonna in datetime
                df.dropna(subset='Time',inplace=True) #tolgo tutte le righe che sono NaN nella colonna time
                
                df[df.columns[1:7]] = df[df.columns[1:7]].apply(pd.to_numeric, errors='coerce')

                #identifico se la lunghezza della traccia può essere divisa perfettamente per il numero di tracce desiderate
                n = (len(df)-1)%ntrack
                if n == 0: #se può essere divisa perfettamente
                    sep = int((len(df)-1)/ntrack)
                else: #se non può essere divisa perfettamente
                    df.drop(arange(n),inplace=True) #tolgo gli ultimi valori utili per avere una divisibilità buona
                    df = df.reset_index(drop=True)
                    sep = int((len(df)-1)/ntrack)
                
                # print(sep) #for debug

                # Aggiunta valori alle liste <=============================
                

                ntrack_id_letter = 1 #variabile che mi tiene conto del numero di traccia di una misura (es: D1 track 1, D1 trak 2, D1 track 3, ...)
                for i in range(ntrack):
                    fileIDs.append(file) #salvo il nome del file 
                    letter_IDs.append(letter_ID + str(letter_id_number)) #salvo la lettera del file
                    ntrack_id.append(ntrack_id_letter)
                    ntrack_id_letter += 1
                    
                    # print(letter_ID) # for debug
                    # print('##',i, i+i*sep) #for debug

                    # Inizio e Fine
                    inizio.append(df['Time'][i + i * sep].time()) #salvo la fine
                    fine.append(df['Time'][(i+1)*sep].time()) #salvo l'inizio
                    
                    # Durata
                    dT = df['Time'][(i+1)*sep] - df['Time'][i + i * sep]
                    dT_str = f"{dT.components[1]:02d}:{dT.components[2]:02d}:{dT.components[3]:02d}"
                    durata.append(dT_str)

                    
                    #LeqA
                    LeqA_min.append(round(min(df['LAeq'][0+i*int(sep):(i+1)*int(sep)]),decimals))
                    LeqA_max.append(round(max(df['LAeq'][0+i*int(sep):(i+1)*int(sep)]),decimals))
                    LeqA_eq.append(round(10*log10(sum(10**(df['LAeq'][0+i*int(sep):(i+1)*int(sep)]/10))/(len(df['LAeq'][0+i*int(sep):(i+1)*int(sep)])) ),decimals))
                    # print(len(df['LAeq'][0+i*int(sep):(i+1)*int(sep)]))
                    # input()

                    #LeqC
                    LeqC_min.append(round(min(df['LCeq'][0+i*int(sep):(i+1)*int(sep)]),decimals))
                    LeqC_max.append(round(max(df['LCeq'][0+i*int(sep):(i+1)*int(sep)]),decimals))
                    LeqC_eq.append(round( 10*log10(sum(10**(df['LCeq'][0+i*int(sep):(i+1)*int(sep)]/10))/(len(df['LCeq'][0+i*int(sep):(i+1)*int(sep)])) ),decimals))

                    PeakC_max.append(round(max(df['LCpeak'][0+i*int(sep):(i+1)*int(sep)]),decimals))
                    PeakC_eq.append(round(average(df['LCpeak'][0+i*int(sep):(i+1)*int(sep)]),decimals))

                    #LAeqS and LAeqI
                    LASeq_T.append(round(10*log10(sum(10**(df['LASeqT'][0+i*int(sep):(i+1)*int(sep)]/10))/(len(df['LASeqT'][0+i*int(sep):(i+1)*int(sep)]))),decimals))
                    LAIeq_T.append(round(10*log10(sum(10**(df['LAIeqT'][0+i*int(sep):(i+1)*int(sep)]/10))/(len(df['LAIeqT'][0+i*int(sep):(i+1)*int(sep)]))),decimals))

                letter_id_number += 1
            
            df_tot['fileID'] = fileIDs
            df_tot['letter_ID'] = letter_IDs
            df_tot['nTrack'] = ntrack_id
            df_tot['inizio'] = inizio
            df_tot['fine'] = fine
            df_tot['durata'] = durata
            df_tot['LeqA_min'] = LeqA_min
            df_tot['LeqA_max'] = LeqA_max
            df_tot['LeqA_eq'] = LeqA_eq 
            df_tot['LASeq_T'] = LASeq_T
            df_tot['LAIeq_T'] = LAIeq_T
            df_tot['LeqC_min'] = LeqC_min
            df_tot['LeqC_max'] = LeqC_max
            df_tot['LeqC_eq'] = LeqC_eq
            df_tot['PeakC_max'] = PeakC_max
            df_tot['PeakC_eq'] = PeakC_eq
            
            

            return df_tot

        elif format == 'xlsx':
            from numpy import average, min, max
            from glob import glob
            from config import SHEET_NAME_XLSX

            
            print('Reading xlsx files')
            #inizializzo il dataFrame pandas con il riassunto delle misure
            df_tot = pd.DataFrame(columns=['fileID','letter_ID','nTrack','LeqA_min','LeqA_max','LeqA_eq','LeqC_min','LeqC_max','LeqC_eq','PeakC_max','PeakC_eq','LASeq_T','durata','inizio','fine'])
            
            #Inizializzazione delle variabili di df_tot
            fileIDs = []
            letter_IDs = []
            inizio = []
            fine = []
            durata = []
            ntrack_id = []
            LeqA_min = []
            LeqA_max = []
            LeqA_eq = []
            LeqC_min = []
            LeqC_max = []
            LeqC_eq = []
            PeakC_max = []
            PeakC_eq = []
            LASeq_T = []

            #lista dei file che terminano con xlsx
            files = glob('*.xlsx')
            files.sort() #riordino i valori della lista in ordine crescente di orario
            
            #itero su tutti i file per estrarre i dati
            letter_id_number = 1 #inizializzo il numero della misura ad 1 per ogni nuovo file misure ( in modo che d1,d2,d3...f1,f2,f3...)
            for file in files:
                # Carico e ripulisco il dataframe
                df = pd.read_excel(file, sheet_name=SHEET_NAME_XLSX) #leggo il file
                df.iloc[:,1] = pd.to_datetime(df.iloc[:,1], format='%Y-%m-%d  %H:%M:%S', errors='coerce')
                df.dropna(subset=df.columns[2],inplace=True) #tolgo tutte le righe che sono NaN nella colonna LAeq
                
                
                #identifico se la lunghezza della traccia può essere divisa perfettamente per il numero di tracce desiderate
                n = (len(df)-1)%ntrack
                if n == 0: #se può essere divisa perfettamente
                    sep = int((len(df)-1)/ntrack)
                else: #se non può essere divisa perfettamente
                    df.drop(arange(n),inplace=True) #tolgo gli ultimi valori utili per avere una divisibilità buona
                    df = df.reset_index(drop=True)
                    sep = int((len(df)-1)/ntrack)
                
                
                # Popolo il dataframe totale
                ntrack_id_letter = 1 #variabile che mi tiene conto del numero di traccia di una misura (es: D1 track 1, D1 trak 2, D1 track 3, ...)
                for i in range(ntrack):
                    fileIDs.append(file) #salvo il nome del file 
                    letter_IDs.append(letter_ID + str(letter_id_number)) #salvo la lettera del file
                    ntrack_id.append(ntrack_id_letter)
                    ntrack_id_letter += 1

                    inizio.append(df[df.columns[1]][i + i * sep].time()) 
                    fine.append(df[df.columns[1]][(i+1)*sep].time())
                    
                    # Durata
                    dT = df[df.columns[1]][(i+1)*sep] - df[df.columns[1]][i + i * sep]
                    dT_str = f"{dT.components[1]:02d}:{dT.components[2]:02d}:{dT.components[3]:02d}"
                    durata.append(dT_str)


                    #LeqA
                    LeqA_min.append(round(min(df[df.columns[2]][0+i*int(sep):(i+1)*int(sep)]),decimals))
                    LeqA_max.append(round(max(df[df.columns[2]][0+i*int(sep):(i+1)*int(sep)]),decimals))
                    LeqA_eq.append(round(10*log10(sum(10**(df[df.columns[2]][0+i*int(sep):(i+1)*int(sep)]/10))/(len(df[df.columns[2]][0+i*int(sep):(i+1)*int(sep)])) ),decimals))
                    # print(len(df['LAeq'][0+i*int(sep):(i+1)*int(sep)]))
                    

                    #LeqC
                    LeqC_min.append(round(min(df[df.columns[4]][0+i*int(sep):(i+1)*int(sep)]),decimals))
                    LeqC_max.append(round(max(df[df.columns[4]][0+i*int(sep):(i+1)*int(sep)]),decimals))
                    LeqC_eq.append(round( 10*log10(sum(10**(df[df.columns[4]][0+i*int(sep):(i+1)*int(sep)]/10))/(len(df[df.columns[4]][0+i*int(sep):(i+1)*int(sep)])) ),decimals))

                    PeakC_max.append(round(max(df[df.columns[3]][0+i*int(sep):(i+1)*int(sep)]),decimals))
                    PeakC_eq.append(round(average(df[df.columns[3]][0+i*int(sep):(i+1)*int(sep)]),decimals))

                    #LAeqS and LAeqI
                    LASeq_T.append(round(10*log10(sum(10**(df[df.columns[5]][0+i*int(sep):(i+1)*int(sep)]/10))/(len(df[df.columns[5]][0+i*int(sep):(i+1)*int(sep)]))),decimals))

                letter_id_number += 1
            
            df_tot['fileID'] = fileIDs
            df_tot['letter_ID'] = letter_IDs
            df_tot['nTrack'] = ntrack_id
            df_tot['inizio'] = inizio
            df_tot['fine'] = fine
            df_tot['durata'] = durata
            df_tot['LeqA_min'] = LeqA_min
            df_tot['LeqA_max'] = LeqA_max
            df_tot['LeqA_eq'] = LeqA_eq 
            df_tot['LASeq_T'] = LASeq_T
            df_tot['LeqC_min'] = LeqC_min
            df_tot['LeqC_max'] = LeqC_max
            df_tot['LeqC_eq'] = LeqC_eq
            df_tot['PeakC_max'] = PeakC_max
            df_tot['PeakC_eq'] = PeakC_eq

            return df_tot







        else:
            error("Sono supportati solo formati 'txt' o 'csv' nella voce format")


    def write_csv(df,file):
        '''
        Funzione che scrive il dataframe df in un file csv
        '''
        df.to_csv(file)
        print('Salvataggio in csv del file completato')


    def write_exel(df, file):
        '''
        Funzione che scrive il dataframe df df in un file excel
        '''
        df.to_excel(file, index=False)
        print('Salvataggio del file exel completato')


    def write_excel_append(df, file, sheet_name, mode_writer = 'w'):
        with pd.ExcelWriter(file, mode = mode_writer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name = sheet_name, index=False, header= True)

    def get_scheda_DPI(excel_info_dir, name_excel_info, sheet_name='Scheda_DPI'):
        '''
        Funzione che legge il foglio "Scheda_DPI" del file excel informativo e restituisce
        un dataframe pandas con i dati dei dispositivi di protezione individuale (DPI).

        INPUT:
            excel_info_dir   = <str>, directory in cui è salvato il file excel informativo
            name_excel_info  = <str>, nome del file excel informativo (es. 'scheda_gruppi_dpi.xlsx')
            sheet_name       = <str>, nome del foglio da leggere (default: 'Scheda_DPI')

        OUTPUT:
            df_dpi = <pd.DataFrame>, dataframe con le seguenti colonne:
                - codice_dpi  : codice identificativo del DPI (str)
                - descrizione : descrizione del DPI (str)
                - marca       : marca del DPI (str)
                - modello     : modello del DPI (str)
                - SNR         : valore SNR del DPI (float)
                - H           : valore di attenuazione alle alte frequenze (float)
                - L           : valore di attenuazione alle basse frequenze (float)
                - M           : valore di attenuazione alle medie frequenze (float)
                - beta        : coefficiente correttivo per l'attenuazione reale (float)
        '''
        import os
        import warnings

        file_path = os.path.join(excel_info_dir, name_excel_info)

        try:
            # Leggo il foglio saltando la riga del titolo (riga 1),
            # usando la riga 2 come intestazione e i dati dalla riga 3 in avanti
            df_raw = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                skiprows=1,   # salta la riga del titolo "Tabella DPI"
                header=0      # usa la riga successiva come intestazione
            )

            # Rinomino le colonne secondo la convenzione richiesta
            col_map = {
                df_raw.columns[0]: 'codice_dpi',
                df_raw.columns[1]: 'descrizione',
                df_raw.columns[2]: 'marca',
                df_raw.columns[3]: 'modello',
                df_raw.columns[4]: 'SNR',
                df_raw.columns[5]: 'H',
                df_raw.columns[6]: 'L',
                df_raw.columns[7]: 'M',
                df_raw.columns[8]: 'beta',
            }
            df_dpi = df_raw.rename(columns=col_map)

            # Forzo i tipi di dato
            str_cols   = ['codice_dpi', 'descrizione', 'marca', 'modello']
            float_cols = ['SNR', 'H', 'L', 'M', 'beta']

            for col in str_cols:
                df_dpi[col] = df_dpi[col].astype(str).replace('nan', None)

            for col in float_cols:
                df_dpi[col] = pd.to_numeric(df_dpi[col], errors='coerce')

            # Messaggio di conferma con elenco marca-modello
            print('Lettura scheda DPI avvenuta con successo. Verifica la correttezza dei dati:')
            for _, row in df_dpi.iterrows():
                print(f'  [{row["codice_dpi"]}]  Marca: {row["marca"]}  -  Modello: {row["modello"]}')

            return df_dpi

        except Exception as e:
            warnings.warn(f'Lettura file excel non avvenuta con successo. Errore: {e}')
            return None



    def read_csv(file):
        '''
        Funzione che legge il file csv grezzo e lo trasforma in un dataFrame Pandas
        '''
        df = pd.read_csv(file) 

        return df


class manager:
    '''
    Classe che gestisce la lettura e la creazione dei file.
    '''
    
    def __init__(self):
        '''
        Salvataggio delle cartelle di misura
        '''

        print(''''
        #############################
            Analisi VR Rumore
        #############################
        ''')

        from os import getcwd, listdir
        self.main_dir = getcwd() #salva la directory principale di lavoro
        print(f'Directory di lavoro: {self.main_dir}') #stampa la directory di lavoro

        self.file_list = listdir(self.main_dir) #salva la lista dei file nella directory principale
        self.file_list.sort() #riordino i file in ordine crescente

        # print(self.file_list)
        if ".DS_Store" in self.file_list:
            self.file_list.remove(".DS_Store")
        
        if "data" in self.file_list:
            self.file_list.remove("data")
        
        if "VR_8h.csv" in self.file_list and "VR_8h.xlsx" in self.file_list:
            self.file_list.remove("VR_8h.csv")
            self.file_list.remove("VR_8h.xlsx")
        
        if "VR_rumore.out" in self.file_list:
            self.file_list.remove("VR_rumore.out")

        
        print(f'Lista dei file nella directory: {self.file_list}') # stampa la lista dei file nella directory principale


    def iterate_directory(self,file_name = 'LSOURCES.txt', format = 'txt', versione_lettura='1'):
        '''
        Funzione che itera su tutte le directory per salvare i file.
        Crea i file csv e xlsx e colora le colonne opportune del file xlsx mediando tutti i dati.
        INPUT: 
            file_name = <str>, default: LSOURCES.txt. Nome dei file di input.
            format = <str>, formato file. Disponibili: 'txt' o 'csv'
        '''
        from os import path, mkdir
        from config import READING_TXT_DATA_FOLDER_NAME, READING_CSV_DATA_FOLDER_NAME, READING_EXEL_DATA_FOLDER_NAME
        
        def read_W_file_txt(file: str, letter_ID: str, decimals: int = 1) -> pd.DataFrame:
            """
            Legge un file in formato 'dati.txt' (output strumento WED con 6 sorgenti,
            separatore decimale virgola, etichette 'Fast A' / 'Fast C' / 'Picco C')
            e restituisce un DataFrame pandas compatibile con l'output di read_measure_file()
            nel formato txt classico (LSOURCES.txt).

            INPUT:
                file       : <str>  percorso del file da leggere
                letter_ID  : <str>  lettera identificativa della sessione (es. 'D', 'F', ...)
                decimals   : <int>  cifre decimali per l'arrotondamento (default 1)

            OUTPUT:
                df : pd.DataFrame con colonne:
                    fileID, letter_ID, nTrack,
                    LeqA_min, LeqA_max, LeqA_eq,
                    LeqC_min, LeqC_max, LeqC_eq,
                    PeakC_max, PeakC_eq,
                    durata, inizio, fine
            """

            # ── lettura file ─────────────────────────────────────────────────────────
            with open(file, 'r', encoding='utf-16') as f:
                lines = f.readlines()

            # ── inizializzazione colonne output ──────────────────────────────────────
            fileIDs    = []
            letter_IDs = []
            nTrack_col = []
            LeqA_min   = []
            LeqA_max   = []
            LeqA_eq    = []
            LeqC_min   = []
            LeqC_max   = []
            LeqC_eq    = []
            PeakC_max  = []
            PeakC_eq   = []
            durata_col = []
            inizio_col = []
            fine_col   = []

            # ── helper: converte stringa con virgola decimale → float ────────────────
            def to_float(s: str) -> float:
                return float(s.strip().replace(',', '.'))

            # ── helper: estrae i valori di una sorgente da una riga tabulare ─────────
            # Ogni sorgente occupa 4 campi: [Leq_eq, Lmin, Lmax, durata]
            # (per Picco C: [vuoto, Lmin, Lmax, durata] → Lmin=PeakC_min skip, Lmax=PeakC_max)
            # La riga inizia con l'etichetta di ubicazione, poi i campi per tutte le sorgenti.
            def parse_row(tokens: list, n_sorgenti: int):
                """
                Restituisce lista di dict con chiavi eq, min, max, dur per ogni sorgente.
                Per 'Picco C' il campo 'eq' sarà vuoto-stringa (lo gestiamo dopo).
                """
                result = []
                # tokens[0] = etichetta ubicazione, poi gruppi di 4 per ogni sorgente
                for i in range(n_sorgenti):
                    base = 1 + i * 4
                    result.append({
                        'eq':  tokens[base].strip(),
                        'min': tokens[base + 1].strip(),
                        'max': tokens[base + 2].strip(),
                        'dur': tokens[base + 3].strip(),
                    })
                return result

            # ── scansione linee ───────────────────────────────────────────────────────
            file_counter = 0  # contatore file nella sessione (→ letter_ID + numero)

            i = 0
            while i < len(lines):
                line = lines[i]

                if line.startswith('File\t'):
                    # ── nuova misurazione ─────────────────────────────────────────────
                    file_counter += 1
                    fileID = line.split('\t')[1].strip()

                    # Inizio e Fine (righe +1, +2)
                    inizio_str = lines[i + 1].split('\t')[1].strip()   # es. "24/09/2025 09:43:24"
                    fine_str   = lines[i + 2].split('\t')[1].strip()

                    # Numero di sorgenti dalla riga +3  →  "Sorgente\t1\t2\t3\t4\t5\t6"
                    sorgente_tokens = lines[i + 3].strip().split('\t')
                    n_sorgenti = len(sorgente_tokens) - 1  # esclude la parola 'Sorgente'

                    # Righe dati: +7 = Fast A, +8 = Fast C, +9 = Picco C
                    row_fastA  = lines[i + 7].strip().split('\t')
                    row_fastC  = lines[i + 8].strip().split('\t')
                    row_piccoC = lines[i + 9].strip().split('\t')

                    vals_A = parse_row(row_fastA,  n_sorgenti)
                    vals_C = parse_row(row_fastC,  n_sorgenti)
                    vals_P = parse_row(row_piccoC, n_sorgenti)

                    # ── una riga per ogni sorgente ────────────────────────────────────
                    for s in range(n_sorgenti):
                        fileIDs.append(fileID)
                        letter_IDs.append(letter_ID + str(file_counter))
                        nTrack_col.append(s + 1)

                        # LeqA
                        LeqA_eq .append(round(to_float(vals_A[s]['eq']),  decimals))
                        LeqA_min.append(round(to_float(vals_A[s]['min']), decimals))
                        LeqA_max.append(round(to_float(vals_A[s]['max']), decimals))

                        # LeqC
                        LeqC_eq .append(round(to_float(vals_C[s]['eq']),  decimals))
                        LeqC_min.append(round(to_float(vals_C[s]['min']), decimals))
                        LeqC_max.append(round(to_float(vals_C[s]['max']), decimals))

                        # Picco C  →  'eq' è vuoto nel file; usiamo min=PeakC_min (non salvato),
                        #             max=PeakC_max; PeakC_eq calcolato come media aritmetica
                        #             (per coerenza con il codice originale che usa average())
                        PeakC_max.append(round(to_float(vals_P[s]['max']), decimals))
                        # PeakC_eq: nel formato LSOURCES era calcolato come media dei picchi
                        # istantanei; qui abbiamo solo il valore complessivo → lo usiamo come eq
                        PeakC_eq .append(round(to_float(vals_P[s]['min']), decimals))  # 'min' = Lmin del picco

                        # durata, inizio, fine
                        durata_col.append(vals_A[s]['dur'])
                        inizio_col.append(inizio_str)
                        fine_col  .append(fine_str)

                    i += 10  # salta l'intero blocco (10 righe per blocco)
                    continue

                i += 1

            # ── costruzione DataFrame ─────────────────────────────────────────────────
            df = pd.DataFrame({
                'fileID'    : fileIDs,
                'letter_ID' : letter_IDs,
                'nTrack'    : nTrack_col,
                'LeqA_min'  : LeqA_min,
                'LeqA_max'  : LeqA_max,
                'LeqA_eq'   : LeqA_eq,
                'LeqC_min'  : LeqC_min,
                'LeqC_max'  : LeqC_max,
                'LeqC_eq'   : LeqC_eq,
                'PeakC_max' : PeakC_max,
                'PeakC_eq'  : PeakC_eq,
                'durata'    : durata_col,
                'inizio'    : inizio_col,
                'fine'      : fine_col,
            })

            print(f'Lettura completata: {file_counter} file, {len(df)} righe totali ({n_sorgenti} sorgenti per file)')
            return df


        #Verifica se esiste la cartella '/data' ed in caso la crea
        self.out_file_dir = self.main_dir + '/' + 'data'
        if path.isdir(self.out_file_dir) == False:
                    mkdir(self.out_file_dir)


        # Caso TXT ==============================
        if format=='txt':
            print('! TXT mode !')
            

            for dir in self.file_list:
                print(f'Iterazione sulla directory: {dir}') 
                chdir(self.main_dir + '/' + dir + '/' + dir) #entro nella cartella della misura i-esima
            
                if dir in READING_TXT_DATA_FOLDER_NAME:
                    df = read_W_file_txt('misW.txt','W')
                else:
                    df, num_of_track, n_files = files.read_measure_file(file_name,list(dir)[-1], format = format)
                
                files.write_csv(df, f'{self.out_file_dir}/{dir}.csv')
                files.write_exel(df, f'{self.out_file_dir}/{dir}.xlsx')
                exel_file.adjust_column_lenght(f'{self.out_file_dir}/{dir}.xlsx', ['A'])
                exel_file.color_column(f'{self.out_file_dir}/{dir}.xlsx', ['F','I','K'], ['FFFF00','FFFF00','FFFF00'])
       
        # Caso CSV ==============================
        elif format=='csv':
            print('! CSV mode !')
            
            
            for dir in self.file_list:
                print(f'Iterazione sulla directory: {dir}')    
                chdir(self.main_dir + '/' + dir) # entro nella cartella dei dati (misE, misF, ecc...)
                
                if dir in READING_TXT_DATA_FOLDER_NAME:                
                    df = read_W_file_txt('misW.txt','W')

                elif dir in READING_EXEL_DATA_FOLDER_NAME:
                    df = files.read_measure_file(file_name, letter_ID=list(dir)[-1], format='xlsx',read_version=versione_lettura) #salvo i DF totale delle misure (A,B,C,G,H)

                else:
                    df = files.read_measure_file(file_name,letter_ID=list(dir)[-1], format='csv', read_version=versione_lettura) # salvo il DF totale delle misure  (D,E,F,ecc)
                
                #salvo i file nella cartella opportuna e nelle versioni csv ed exel
                files.write_csv(df, f'{self.out_file_dir}/{dir}.csv')
                files.write_exel(df, f'{self.out_file_dir}/{dir}.xlsx')
                exel_file.adjust_column_lenght(f'{self.out_file_dir}/{dir}.xlsx', ['A'])
                exel_file.color_column(f'{self.out_file_dir}/{dir}.xlsx', ['F','I','J'], ['FFFF00','FFFF00','FFFF00'])






        
        
        else:
            error("Disponibili solo i formati 'csv' o 'txt'. Controlla di aver inserito il testo correttamente.")


    
    def VR8h_exel(self, name_averaged_data, out_VR8h_name):
        '''
        Funzione che gestisce il file finale della valutazione del rischio in 8h dei lavori. 
        Importa il file xlsx e ne colora le colonne del colore specifico in base al grado di rischio.
        '''
        a = analisi(self.main_dir)
        a.VR_8h(name_averaged_data)
        exel_file.color_cell_VR8h(out_VR8h_name)

    


class analisi:
    '''
    Classe con i metodi per l'analisi delle misure
    '''

    def __init__(self, csv_data_directory):
        '''
        INPUT:
            df = <pd.DataFrame>, dataframe con i dati della misura
        '''
        self.main_dir = csv_data_directory #salvataggio della directory principale
        


    # prende i valori raw e crea un 
    def average_values(self):
        '''
        Funzione che concatena i dataframe di tutte le misure e scrive un file completo con le medie
        e le std sulla media (incertezze)
        '''
        import glob 
        from os.path import exists

        if not exists(self.main_dir + '/averaged_data.csv'):
            

            files_csv = glob.glob(self.main_dir + '/*.csv') #leggo solo i file con estensione csv
            files_csv.sort()
            # print(files)

            df_list = [] # lista dei dataframe che ci sono
            
            #inizializzo il dataFrame pandas con i valori medi di tutte le misure
            df_avg = pd.DataFrame(columns=['jobName', 'ID', 'U' ,'LeqA','LeqC','Ppeak'])


            # salvo in un dataframe tutti i file csv
            for file in files_csv:
                # print(file)
                df_list.append(pd.read_csv(file))
            
            # iterazione su tutti i dataframe
            for df in df_list:
                
                fileIDs = df[df.columns[1]].unique() #lista dei fileID
                letter_IDs = df[df.columns[2]].unique() #lista dei letterID
                
                LeqA_mean = zeros(len(fileIDs)) # inizializzo l'array di LeqA_mean
                LeqC_mean = zeros(len(fileIDs)) # inizializzo l'array di LeqC_mean
                Ppeak_mean = zeros(len(fileIDs)) # inizializzo l'array di LeqC_mean
                U_sdom = zeros(len(fileIDs)) #standard deviation of mean (incertezza misure)
                

                for i in range(len(fileIDs)):
                    idx = df[df.columns[1]] == fileIDs[i] # prendo solo i valori opportuni
                    
                    #calcolo i valori medi
                    LeqA_mean[i] = round(mean(df['LeqA_eq'][idx]),1)
                    LeqC_mean[i] = round(mean(df['LeqC_eq'][idx]),1) # + std(df['LeqC_max'][idx],ddof=1) (in caso da aggiungere se serve)
                    Ppeak_mean[i] = round(max(df['PeakC_max'][idx]),1) # + 1.56 in caso da aggiungere 
                    # print(LeqA_mean)
                    
                    #calcolo l'incertezza sulla misura LeqA (SDOM)
                    U_sdom[i] = round(std(df['LeqA_max'][idx], ddof=1) * sqrt(1/sum(idx)),1)

                
                new_df = pd.DataFrame({'jobName': fileIDs ,
                                       'ID':letter_IDs ,
                                       'U':U_sdom,
                                       'LeqA' : LeqA_mean, 
                                       'LeqC' : LeqC_mean ,
                                       'Ppeak' : Ppeak_mean})
                df_avg = pd.concat([df_avg,new_df], ignore_index=True)
            
            
            df_avg['Ti'] = [[]] * len(df_avg) #creo la colonna con i valori di exposure time
            df_avg['GrOm'] = [[]] * len(df_avg) #creo la colonna con gli ID del gruppo omogeneo
            # df_avg['DPI'] = [[]] * len(df_avg) #creo la colonna contenente il riferimento all'uso del DPI


            files.write_csv(df_avg, self.main_dir + '/averaged_data.csv')
            files.write_exel(df_avg, self.main_dir + '/averaged_data.xlsx')
            exel_file.adjust_column_lenght(self.main_dir + '/averaged_data.xlsx', 'A')
            print('Averaged data files created')
            return df_avg
        else:
            print('File averaged.csv already exists!')
            df_avg = pd.read_csv(self.main_dir + '/averaged_data.csv')
            return df_avg

    
    def get_scheda_info(self, df_avg, excel_info_dir=None, name_exel_info="scheda_gruppi_dpi.xlsx"):
        '''
        Funzione che prende in ingresso df_avg e il file excel con le schede mansioni e ne fa il merge con la corrispondenza
        dell'ID misura. 
        Unisce i dataframe in un unico dataframe chiamato df_global e lo salva in /data.
        
        OUTPUT:
            df_global = pd.dataFrame, contiene l'unione dell'excel scheda_gruppi_dpi.xlsx e df_avg.
        '''
        import pandas as pd
        from os.path import dirname, exists
        from config import SCHEDA_MANSIONI


        # ------------------------------------------------------------------
        # Variabili interne (modificabili facilmente)
        # ------------------------------------------------------------------
        sheet_mansioni = SCHEDA_MANSIONI   # nome foglio scheda mansioni

        # Percorso default del file excel: una directory sopra self.main_dir
        # (corrisponde a main_directory, dato che self.main_dir = main_directory/data)
        if excel_info_dir is None:
            excel_info_dir = dirname(self.main_dir) + ''


        # Percorsi di output (scrittura del df_global in formato excel sovrascrivendo il precedente)
        base_name, ext = name_exel_info.rsplit('.', 1)
        out_xlsx = f"{excel_info_dir}/{base_name}_mod.{ext}"

        # ------------------------------------------------------------------
        # Verifica esistenza file excel
        # ------------------------------------------------------------------
        if not exists(excel_info_dir):
            raise FileNotFoundError(
                f"File excel non trovato: '{excel_info_dir}'\n"
                f"Controlla che il file {name_exel_info} sia presente in: "
                f"'{dirname(excel_info_dir)}' e si chiami {name_exel_info}"
            )

        print(f'Lettura file scheda gruppi: {excel_info_dir}')
        
        

        # ==================================================================
        # STEP 1 — Lettura foglio 'Scheda_mansioni' come dataframe
        # ==================================================================
        df_scheda_gruppi_dpi = pd.read_excel(excel_info_dir + "/" + name_exel_info, sheet_name=sheet_mansioni, header=1)

        df_scheda_gruppi_dpi = df_scheda_gruppi_dpi.merge(
            df_avg[['ID', 'U', 'LeqA', 'LeqC', 'Ppeak']],
            left_on='ID_misura',
            right_on='ID',
            how='left'
        ).drop(columns=['ID'])


        df_avg = df_avg.merge(
            df_scheda_gruppi_dpi[['ID_misura']],
            left_on = 'ID',
            right_on = 'ID_misura',
            how='left',
            indicator=True
        )

        # Prendo eventuali valori che non sono stati usati da df_avg
        missing_ids = df_avg.loc[df_avg['_merge'] == 'left_only', 'ID'].unique().tolist()

        

        if missing_ids:
            print(
                f'\n*** WARNING: Le seguenti misure di df_avg non sono state usate:\n '
                f'    Misure mancanti: {missing_ids}\n'
                f'***'
            )

        # ==================================================================
        # STEP 4 — Salvataggio (sovrascrittura averaged_data.csv e .xlsx)
        # ==================================================================
        # files.write_csv(df_avg, out_csv)


        files.write_excel_append(df_scheda_gruppi_dpi, out_xlsx, sheet_mansioni )
        print(f'\nFile aggiornati sovrascritti:\n  XLSX -> {out_xlsx}')

        return df_scheda_gruppi_dpi

    #DEPRECATED, sostituita da analisi_8h()
    def calcolo_Leq8h(self, df_GrOm, T0 = 480):
        '''
        funzione che calcola il livello equivalente nelle 8h
        INPUT:
            df_GrOm = <dataFrame>, contenente i valori medi delle misure di uno specifico gruppo omogeneo
            T0 = <int>, numero di minuti di esposizione di una giornata lavorativa
        OUTPUT:
            df_avg = <dataFrame> specifico della mansione
        
        
        Theory:
        livello sonoro equivalente 8h
        Lex,8h = 10*log( 1/T0 * sum( T_i * 10^(0.1 * L_i)  )    (dBA)
            
            La i indica la sorgente sonora i_esima
            T0 è il tempo totale di lavoro in ore (8 h lavorative in genere)
            T_i è il tempo di esposizione quotidiana, in ore, di un lavoratore alla fonte i-esima
            L_i è il livello equivalente continuo ponderato A della fonte i-esima
        '''

        # Calcolo il Leq_8h
        self.LeqA_8h = 10 * log10( 1/T0 * dot( df_GrOm['Ti'], 10**(0.1 * df_GrOm['LeqA']) ) )
        return self.LeqA_8h
    

    def calcolo_U_estesa(self, df_GrOm, T0 = 480, u2m = 0.7, u_pos = 1):
        '''
        Funzione che calcola l'incertezza combinata standard e quella estesa per il LeqA_8h
        INPUT:
            df_GrOm = <dataFrame>, contenente i valori medi delle misure di uno specifico gruppo omogeneo
            T0 = <int>, minuti di esposizione di una giornata lavorativa, in genere 8 ore ossia 480 minuti
            u2m = <float>, errore secondo normativa in base allo strumento (0.7 o 1.5)
            u_pos = <float>, errore nel posizionamento dello strumento (in metri)
        OUTPUT:
            U_estesa = <float>, incertezza estesa
            U_comb_std = <float>, iincertezza combinata standard

        THEORY:

        '''
        LeqA8H = self.calcolo_Leq8h(df_GrOm, T0=T0)
        self.U_comb_std = sum(  (df_GrOm['Ti']/T0  * 10**(0.1*( df_GrOm['LeqA'] - LeqA8H ) ) )**2   * ( df_GrOm['U']**2 + u2m**2 + u_pos **2 ) +
                    ( 4.34 * (1/T0  * 10**(0.1*( df_GrOm['LeqA'] - LeqA8H ) )) * (std(df_GrOm['Ti'], ddof=1) * sqrt(1/len(df_GrOm)) ) )**2 )
        
        self.U_ext = self.U_comb_std * 1.65

        return self.U_ext, self.U_comb_std


    def analisi_8h(self, output_dir, df_HEG: pd.DataFrame, T0=480, u2m=0.7, u_pos=1.0):
        '''
        Funzione che calcola la valutazione del rischio rumore su base giornaliera (8h)
        per ogni gruppo omogeneo presente in df_HEG.

        Per ogni gruppo omogeneo vengono calcolati:
            - Lex8h    : livello sonoro equivalente ponderato A su 8h  
            - U        : incertezza estesa                             
            - Lex_max  : Lex8h + U                                     
            - L_picco_C: massimo dei Ppeak nel gruppo omogeneo         

        TEORIA (D.Lgs. 81/08, norma ISO 9612):

            AG_i    = Ti/T0 * 10^(LeqA_i / 10)              
            Lex8h   = 10 * log10( SUM(AG_i) )               

            Z_i     = Ti/T0 * 10^((LeqA_i - Lex8h) / 10)   
            W_i     = max(0, Z_i^2 * (u_i^2 + u2m^2 + u_pos^2)) 
                      (il II termine X_i, legato alla variabilita` di Tm, e` posto = 0)
            U_comb  = SUM(W_i)                               
            U       = 1.65 * sqrt(U_comb)                    

            Lex_max   = Lex8h + U                            
            L_picco_C = max(Ppeak nel gruppo)                

        INPUT:
            output_dir = <str>, directory in cui salvare i file di output
            df_HEG     = <pd.DataFrame>, dataframe con i totali delle misure.
            T0         = <int>, tempo di riferimento in minuti (default 480 = 8h)
            u2m        = <float>, incertezza strumentale U2 (default 0.7 dB)
            u_pos      = <float>, incertezza di posizione U3 (default 1.0 dB)

        OUTPUT (file scritti in output_dir):
            VR8h_totale.csv  e  VR8h_totale.xlsx  : file unico riepilogativo con i risultati di tutti
                                       i gruppi omogenei (GrOm, Lex8h, U, Lex_max, L_picco_C)
            VR8h_<ID_GrOm>_<descrizione_grom>.csv  e  VR8h_<GrOm>_<descrizione_grom>.xlsx  : un file per ogni gruppo omogeneo con
                                       il dettaglio delle misure (ID, LeqA, LeqC, U, Ti)
        '''
        import os
        import pandas as pd
        from config import (NOME_VR8h_totale, 
                            NOME_VR8h_riepilogo,
                            Nome_colonna_IDgrom,
                            Nome_colonna_Descrizione_GrOm,
                            NOME_VR8h_aggiornato)

        # ----------------------------------------------------------------
        # STEP 1 — Lettura unici valori gruppi omogenei
        # ----------------------------------------------------------------
        
        grom_id_unique = df_HEG[Nome_colonna_IDgrom].unique()

        # ----------------------------------------------------------------
        # STEP 2 — Ciclo su ogni gruppo omogeneo: calcoli e raccolta risultati
        # ----------------------------------------------------------------
        summary_rows = []  # raccoglie una riga di riepilogo per ogni gruppo omogeneo del dataframe

        for grom in grom_id_unique:
            
            #Seleziono la restrizione al gruppo omogeneo del dataframe
            df_tmp = df_HEG[df_HEG[Nome_colonna_IDgrom] == grom] 
            mansione = df_tmp[df_tmp[Nome_colonna_IDgrom] == grom][Nome_colonna_Descrizione_GrOm].tolist()[0]

            # STEP 3 — Verifica che la somma dei Ti sia esattamente T0
            tot_ti = df_tmp['Ti'].sum()
            if tot_ti != T0:
                raise ValueError(
                    f"Gruppo omogeneo {grom}| {mansione} : somma dei Ti = {tot_ti} min "
                    f"!= T0 = {T0} min. Controlla i valori di Ti nel file averaged_data."
                )

            leqa  = df_tmp['LeqA'].values
            ti    = df_tmp['Ti'].values
            u_mis = df_tmp['U'].values

            # STEP 4 — Lex8h 
            #          AG_i = Ti/T0 * 10^(LeqA_i / 10)
            lex8h = 10 * log10(sum(ti / T0 * 10**(leqa / 10)))

            # STEP 5 — Incertezza estesa U  
            #          Z_i = Ti/T0 * 10^((LeqA_i - Lex8h) / 10)   [col. Z]
            #          W_i = max(0, Z_i^2 * (u_i^2 + u2m^2 + u_pos^2))  [col. W]
            #          II termine X_i = 0  (Tmax/Tmin non disponibili in df_avg)
            z     = ti / T0 * 10**((leqa - lex8h) / 10)
            w     = max(z**2 * (u_mis**2 + u2m**2 + u_pos**2), 0)
            U_val = 1.65 * sqrt(sum(w))

            # STEP 6 — Lex_max e L_picco_C
            lex_max   = lex8h + U_val
            l_picco_c = max(df_tmp['Ppeak'].values)

            #Calcolo classe di rischio
            if lex_max < 80:
                classe_rischio = 'BASSA'
            elif 80 <= lex_max < 85:
                classe_rischio = 'MEDIA'
            elif lex_max >= 85:
                classe_rischio = 'ALTA'
            else:
                classe_rischio = ""


            # STEP 7 — Accumulo riga di riepilogo
            summary_rows.append({
                'ID_GrOm':   grom,
                'Mansione': mansione,
                'Lex8h':     round(lex8h,     1),
                'U':         round(U_val,     1),
                'Lex_max':   round(lex_max,   1),
                'L_picco_C': round(l_picco_c, 1),
                'classe_rischio': classe_rischio
            })

            # STEP 8 — File di dettaglio per il gruppo: ID, LeqA, LeqC, U, Ti
            total_xlsx_dir = os.path.join(output_dir, NOME_VR8h_totale)

            # Check sull'esistenza della directory
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
                print(f'-- {output_dir} created --')

            if not os.path.exists(total_xlsx_dir):
                with pd.ExcelWriter(total_xlsx_dir, mode='w', engine='openpyxl') as writer:
                    df_tmp.to_excel(writer, sheet_name=f"Scheda {grom}", index=False)
            else:
                with pd.ExcelWriter(total_xlsx_dir, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                    df_tmp.to_excel(writer, sheet_name=f"Scheda {grom}", index=False)

            print('\n\n')
            print(f'Gruppo {grom} | {mansione}: \nLex8h={round(lex8h,1)} dB(A)\n'
                  f'U={round(U_val,1)} dB \nLex_max={round(lex_max,1)} dB(A)\n'
                  f'L_picco_C={round(l_picco_c,1)} dB(C)')


        # ----------------------------------------------------------------
        # STEP 9 — File riepilogativo unico VR8h.csv e VR8h.xlsx
        # ----------------------------------------------------------------
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(os.path.join(output_dir, NOME_VR8h_riepilogo), index=False, sheet_name='Riepilogo')
        


        # Rendiamo belli gli exccel
        exel_file.formatting_excel_VR8h_totale(os.path.join(output_dir,NOME_VR8h_totale),
                                                            os.path.join(output_dir,NOME_VR8h_totale))
        
        
        exel_file.inserisci_valutazione_schede(os.path.join(output_dir,NOME_VR8h_riepilogo),
                                                os.path.join(output_dir,NOME_VR8h_totale),
                                                os.path.join(output_dir, NOME_VR8h_aggiornato))
        
        
        exel_file.colora_classe_rischio(os.path.join(output_dir,NOME_VR8h_riepilogo))


        # Inserisco il riepilogo come primo foglio del file aggiornato
        exel_file.transfer_riepilogo2aggiornato(os.path.join(output_dir,NOME_VR8h_riepilogo),
                                                os.path.join(output_dir,NOME_VR8h_aggiornato))

        # Bordi neri sulle celle popolate
        exel_file.colora_bordi_celle(os.path.join(output_dir,NOME_VR8h_aggiornato))

        print(f'\n###\nanalisi_8h completata.\nDati salvati in {output_dir}\n###')


    def VR_8h(self,df_avg_dir):
            '''
            Funzione che analizza i dati avg per restituire i valori nelle 8h.
            NB: è importante aver modificato il file avergaed_data.xlsx in modo da aver inserito
            il numero identificativo del gruppo omogeneo e il tempo di esposizione della mansione
            INPUT:
                df_avg_dir = <str>, directory del file con la media dei valori
            OUTPUT:
                df_VR8h = <dataFrame>, con i valori di valutazione del rischio in 8h
            '''
            df_avg = files.read_csv(df_avg_dir) #lettura del df in csv

            #creo le variabili vuote del dataframe da inserire 
            LeqA8h = []
            Uext = []
            U_comb_std = []
            Peak_max = []
            LeqA_max = []

            #prendo solo i valori unici degli ID del gruppo omogeneo
            grorIDs = df_avg['GrOm'].unique() # ; print(grorIDs)

            #inizio il ciclo su tutti i valori di ID del gruppo omogeneo
            for i in range(len(grorIDs)):
                print('Calcolo su gruppo omogenero:', grorIDs[i]) 

                idx = df_avg['GrOm'] == grorIDs[i] # prendo solo i valori opportuni di GrOm e gl indici
                
                #calcolo il LeqA su 8 ore
                ax1 = self.calcolo_Leq8h(df_GrOm=df_avg[idx]) 
                LeqA8h.append(ax1)

                #calcolo incertezza estesa
                ax2, ax22 = self.calcolo_U_estesa(df_GrOm=df_avg[idx])
                Uext.append(ax2), U_comb_std.append(ax22)
                
                #Calcolo massimo del picco
                ax3 = max(df_avg['Ppeak'][idx])
                Peak_max.append(ax3)

                #Calcolo il massimo di LeqA
                LeqA_max.append(ax1 + ax2)

                # input('--- pausa ---')
            
            # df_VR8h = pd.DataFrame(columns=['GrOm','LeqA', 'U' , 'Ppeak'])
            df_VR8h = pd.DataFrame({'GrOm': grorIDs,
                                    'LeqA_8h': round(LeqA8h,1),
                                    'Peak': round(Peak_max,1),
                                    'U_ext': round(Uext,1),
                                    'Leq_max': round(LeqA_max,1)})

            files.write_exel(df_VR8h, self.main_dir+ '/VR_8h.xlsx')
            files.write_csv(df_VR8h, self.main_dir+ '/VR_8h.csv')
            print(df_VR8h)
            print('Vautazione rischio 8h completata.')
    

    def DPI_HML(self, VR_8h_dir, H, M, L, beta, grom):
        '''
        Funzione che utilizza il metodo HML per il calcolo dei coefficienti di riduzione dei DPI.
        INPUT:
           VR_8h_dir = <str>, directory del file VR_8h.csv
           H = <float>, valore di attenuazione delle alte frequenze
           M = <float>, valore di attenuazione delle medie freq
           L = <float>, valore di attenuazione delle basse freq
           beta = <foat>, coefficiente correttivo dei DPI
           gror = <list>, lista dei gruppi omogenei su cui applicare questi dpi
        OUTPUT:
            df = <pd.DataFrame>, contenente i valori di LeqA_8h con la riduzione e il valore di PNR

        
        Theory:
            1. Calcolo dL = LeqC - LeqA

            2. Calcolo del PNR

                Se dL <= 2 dB:
                    PNR = M - (H - M)/4 * (dL - 2)
                Se dL > 2:
                    PNR = M - (H - M)/8 * (dL - 2)

            3. Calcolo del livello effettivo all'orecchio
                LeqA_eff = LeqA - PNR

            4. Confronta idonietà del DPI

        '''
        # Aggiunta del coefficiente correttivo 
        H = beta * H
        M = beta * M
        L = beta * L


        df_avg = files.read_csv(self.main_dir + '/averaged_data.csv') #lettura del file csv con le medie dei dati
        df_8h = files.read_csv(VR_8h_dir + '/VR_8h.csv') #lettura del file csv dei dati valutazione rischio 8h

        # grorIDs = df_avg['GrOm'].unique() #scelgo solo i valori unici dei gruppi omogenei
        
        # Inizializzo le nuove colonne del df_avg
        df_avg['PNR'] = 0.0
        df_avg['LeqA_eff'] = 0.0
        #Inizializzo le nuove colonne del df_8h
        df_8h['PNR_avg'] = 0.0
        df_8h['LeqA_8h_eff'] = 0.0

        for i in range(len(grom)): #itero sul numero di gruppi omogenei selezionati
            
            print(f'Calcolo HML dei PDI su gruppo omogeneo {grom[i]}') 

            idx_grom = df_avg['GrOm'] == grom[i] # prendo solo i valori opportuni di GrOm e gli indici
            idx_dpi = idx_grom * df_avg['DPI'] # prendo solo i valori degli indici del gruppo omogeneo in cui compaiono i DPI
            
            dL = df_avg['LeqC'][idx_dpi] - df_avg['LeqA'][idx_dpi] #calcolo il valore dL
            
            # Calcolo del PNR in base al valore di dL con la condizione
            df_avg.loc[idx_dpi, 'PNR'] = round(dL.apply(
                lambda x: M - (H - M) / 4 * (x - 2) if x <= 2 else M - (H - M) / 8 * (x - 2)
            ),1)

            df_8h.loc[df_8h['GrOm']==grom[i], 'PNR_avg'] = mean(df_avg['PNR'][idx_dpi])
            # df_8h.loc[df_8h['GrOm'] == i, 'PNR_avg'] = mean(df_avg['PNR'][idx_dpi][df_avg['PNR'][idx_dpi] != 0.0])


        # AVG
        df_avg['LeqA_eff'] = df_avg['LeqA'] - df_avg['PNR'] # calcolo la colonna con la riduzione di LeqA
        files.write_csv(df_avg,self.main_dir + '/averaged_data.csv') #salvo il file avg in csv
        df_avg['DPI'] = df_avg['DPI'].map({True: 'Si', False: 'No'}) #mappo i valori booleani in si e no per exel
        files.write_exel(df_avg,self.main_dir + '/averaged_data.xlsx') #salvo il file avg in xlsx

        # VR 8h
        df_8h['LeqA_8h_eff'] = round(df_8h['LeqA_8h'] - df_8h['PNR_avg'],1)
        files.write_csv(df_8h,VR_8h_dir + '/VR_8h.csv')
        files.write_exel(df_8h,VR_8h_dir + '/VR_8h.xlsx')
        exel_file.color_cell_VR8h(VR_8h_dir + '/VR_8h.xlsx')


    def applica_DPI_HML(self, excel_info_scheda_dpi, excel_total, excel_output, excel_aggiornato):
        '''
        Funzione che applica il metodo HML per il calcolo dell'attenuazione dei DPI.
        Parte dal file exel totale e calcola l'attenuazione dei dpi. Inserisce poi i valori nell'excel di riepilogo

        Per ogni DPI definito nella scheda e per ogni gruppo omogeneo:
            1. Legge i parametri H, L, M, beta e calcola H' = beta*H, L' = beta*L, M' = beta*M
            2. Calcola diff_C_A = LeqC - LeqA riga per riga
            3. Calcola PNR con metodo HML:
                - se diff_C_A <= 2 : PNR = M' - (H'-M')/4  * (diff_C_A - 2)
                - se diff_C_A >  2 : PNR = M' - (H'-L')/8  * (diff_C_A - 2)
            4. Aggiunge colonne PNR e LeqA_rid = LeqA - PNR
            5. Salva il file (csv + xlsx) nella cartella output_dpi/DPI_i/
            6. Colora le celle di LeqA_rid nel file xlsx in base al livello di rischio:
                < 65            : arancione  (#FF8C00)
                65 <= v <= 70   : giallo sc. (#FFD700)
                75 <= v <= 80   : giallo sc. (#FFD700)
                70 <  v < 75   : verde      (#008000)
                > 80            : rosso      (#DC143C)

        INPUT:
            excel_info_scheda_dpi  = <str>, nome del file excel con la scheda DPI
            excel_total = <str>, path dell'excel totale
            excel_output = <str>, path dell'excel di output (che può anche essere il riepilogo aggiornato)
            excel_aggiornato = <str>, path dell'excel di aggiornamento, quello con le tabelle
        '''
        import glob
        import os
        import re
        import numpy as np
        import openpyxl as ex
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter
        from config import SCHEDA_DPI

        # ── Step 1: leggi la scheda DPI, schede mansioni e output file ──────────────────────────────────────────
        
        #leggi df DPI
        try:
            df_dpi = pd.read_excel(excel_info_scheda_dpi, sheet_name=SCHEDA_DPI, header=1)
            df_dpi['PNR'] = 0.0 #inizializzo a zero la colonna PNR
            df_dpi['LeqA_rid'] = 0.0 #inizializzo a zero la nuova colonna LeqA_rid

        except FileNotFoundError:
            print(f'File {excel_info_scheda_dpi} not found. Check it out')
            

        #  leggi schede mansioni da VR8h totale
        try:
            #prendo la lista degli sheet names da excel_total
            sheet_names_heg = pd.ExcelFile(excel_total).sheet_names 
        except FileNotFoundError:
            print(f'{excel_total} file not found')
        

        #Leggi df output file (excel_riepilogo)
        try:
            df_riepilogo = pd.read_excel(excel_output,header=0)
            
        except FileNotFoundError:
            print(f"{excel_output} not found")




        # ── Step 2: itera sulle schede e sui DPI  ───────────────────────────────
        for idx_sn, sn in enumerate(sheet_names_heg): #itero sulle schede omogenee
            
            #itero sui dpi
            for dpi_idx in range(len(df_dpi)): 
                df_heg = pd.read_excel(excel_total, sheet_name=sn) #leggo il foglio heg e mi salvo i dati
                dpi = df_dpi.loc[dpi_idx] #seleziono la riga del dpi con le informazioni

                print(f'\n── {dpi.codice_DPI} | {dpi.Marca} {dpi.Modello} '
                    f'| β={dpi.Beta}  H={dpi.H:.2f}  M={dpi.M:.2f}  L={dpi.L:.2f} ──')

                # ── Calcolo PNR vettorializzato ───────────────────────────────
                diff_C_A = df_heg['LeqC'] - df_heg['LeqA']

                PNR = np.where(
                    diff_C_A <= 2,
                    dpi.Beta*dpi.M - (dpi.H * dpi.Beta - dpi.M * dpi.Beta) / 4 * (diff_C_A - 2),   # diff <= 2
                    dpi.Beta*dpi.M - (dpi.H * dpi.Beta - dpi.L * dpi.Beta) / 8 * (diff_C_A - 2)    # diff >  2
                )

                PNR = PNR.mean()

                df_dpi.loc[dpi_idx, 'PNR'] = (PNR).round(1) #salvo PNR in nel df dpi
                df_dpi.loc[dpi_idx, 'LeqA_rid'] = (df_riepilogo.loc[idx_sn,'Lex_max'] - PNR).round(1) #savo LeqA_rid

                


                # ── Scrittura su excel_aggiornato ────────────────────────────────────
            
            from openpyxl.styles import Font
            from config import COL_INIZIO_DPI, FIND_TESTO_FINE_TABELLA_VALUTAZIONE, SEPARAZIONE_RIGHE_DA_VALUTAZIONE, TESTO_TITOLO_DPI
            # ── CONFIGURAZIONE ─────────────────────────────────────────────────
            # Colonna di ancoraggio della sezione DPI (14 = colonna N)
            COL_INIZIO = COL_INIZIO_DPI

            # Stringa da cercare in COL_INIZIO per localizzare la fine
            # della tabella VALUTAZIONE SU BASE GIORNALIERA
            TESTO_FINE_VALUTAZIONE = FIND_TESTO_FINE_TABELLA_VALUTAZIONE

            # Righe vuote da lasciare tra fine VALUTAZIONE e titolo sezione DPI
            RIGHE_VUOTE_SEP = SEPARAZIONE_RIGHE_DA_VALUTAZIONE

            # Testo del titolo della nuova sezione
            TESTO_TITOLO_SEZIONE = TESTO_TITOLO_DPI

            # Colonna di df_dpi che occupa più di una cella fisica nel foglio
            NOME_COL_MERGE = 'Marca'
            SPAN_MERGE     = 2          # celle fisiche occupate (es. O+P)

            # Colonna di df_dpi su cui applicare il colore di rischio
            NOME_COL_COLORE = 'LeqA_rid'

            # Soglie colore per NOME_COL_COLORE
            # val < SOGLIA_ARANCIONE_MAX                               → arancione
            # SOGLIA_ARANCIONE_MAX <= val <= SOGLIA_VERDE_MIN           → giallo
            # SOGLIA_VERDE_MIN < val < SOGLIA_VERDE_MAX                 → verde
            # SOGLIA_VERDE_MAX <= val <= SOGLIA_ROSSO_MIN               → giallo
            # val > SOGLIA_ROSSO_MIN                                   → rosso
            SOGLIA_ARANCIONE_MAX = 65
            SOGLIA_VERDE_MIN     = 70
            SOGLIA_VERDE_MAX     = 75
            SOGLIA_ROSSO_MIN     = 80
            COLORE_ARANCIONE = 'FF8C00'
            COLORE_GIALLO    = 'FFD700'
            COLORE_VERDE     = '008000'
            COLORE_ROSSO     = 'DC143C'

            # ── Calcola offset fisici per ogni colonna di df_dpi ──────────────
            # NOME_COL_MERGE occupa SPAN_MERGE celle, tutte le altre 1 cella
            _offset_corrente = 0
            col_offsets = {}   # {nome_colonna_df: offset_intero_da_COL_INIZIO}
            for col in df_dpi.columns:
                col_offsets[col] = _offset_corrente
                _offset_corrente += SPAN_MERGE if col == NOME_COL_MERGE else 1
            larghezza_sezione = _offset_corrente  # celle fisiche totali della sezione

            # ── Apri workbook e seleziona il foglio corrente ──────────────────
            wb_ag = ex.load_workbook(excel_aggiornato)
            ws_ag = wb_ag[sn]

            # ── Trova dinamicamente l'ultima riga della tabella VALUTAZIONE ───
            # Se la cella appartiene a un merge, si usa l'ultima riga del range
            ultima_riga_val = None
            for row_cells in ws_ag.iter_rows(min_col=COL_INIZIO, max_col=COL_INIZIO):
                cell = row_cells[0]
                if cell.value and TESTO_FINE_VALUTAZIONE in str(cell.value):
                    ultima_riga_val = cell.row
                    for mr in ws_ag.merged_cells.ranges:
                        if mr.min_row == cell.row and mr.min_col == COL_INIZIO:
                            ultima_riga_val = mr.max_row
                            break
                    break

            if ultima_riga_val is None:
                print(f"[WARN] '{TESTO_FINE_VALUTAZIONE}' non trovato nel foglio "
                        f"'{sn}': scrittura DPI saltata.")
                wb_ag.close()
            else:
                # ── Calcola righe e colonne di sezione ───────────────────────
                riga_titolo       = ultima_riga_val + RIGHE_VUOTE_SEP + 1
                riga_intestazioni = riga_titolo + 1
                riga_dati_base    = riga_intestazioni + 1
                col_fine_sezione  = COL_INIZIO + larghezza_sezione - 1
                col_merge         = COL_INIZIO + col_offsets[NOME_COL_MERGE]
                col_colore        = COL_INIZIO + col_offsets[NOME_COL_COLORE]
                NOME_COL_VALUTAZIONE = 'Valutazione efficacia'
                col_valutazione   = col_fine_sezione + 1

                # ── Colonna Lpicco_rid: attiva solo se il Massimo dei Lpicco,C ─
                # (riga 4, cercato per etichetta) supera la soglia ──────────────
                NOME_COL_LPICCO_RID = 'Lpicco_rid'
                SOGLIA_LPICCO_MAX = 135
                col_lpicco_rid = col_valutazione + 1

                lpicco_max = None
                for cell in ws_ag[4]:
                    if cell.value and "Massimo dei Lpicco" in str(cell.value):
                        lpicco_max = ws_ag.cell(row=4, column=cell.column + 1).value
                        break

                scrivi_lpicco_rid = isinstance(lpicco_max, (int, float)) and lpicco_max > SOGLIA_LPICCO_MAX
                col_fine_reale = col_lpicco_rid if scrivi_lpicco_rid else col_valutazione

                # ── Titolo sezione: merged su tutta la larghezza, bold ────────
                ws_ag.merge_cells(
                    start_row=riga_titolo,   start_column=COL_INIZIO,
                    end_row=riga_titolo,     end_column=col_fine_reale
                )
                cell_titolo       = ws_ag.cell(row=riga_titolo, column=COL_INIZIO)
                cell_titolo.value = TESTO_TITOLO_SEZIONE
                cell_titolo.font  = Font(bold=True)

                # ── Intestazioni colonne: nomi da df_dpi.columns, bold ────────
                # NOME_COL_MERGE ha il merge su SPAN_MERGE celle
                for nome_col, offset in col_offsets.items():
                    cell_hdr       = ws_ag.cell(row=riga_intestazioni,
                                                column=COL_INIZIO + offset)
                    cell_hdr.value = nome_col
                    cell_hdr.font  = Font(bold=True)
                ws_ag.merge_cells(
                    start_row=riga_intestazioni, start_column=col_merge,
                    end_row=riga_intestazioni,   end_column=col_merge + SPAN_MERGE - 1
                )
                cell_hdr_val       = ws_ag.cell(row=riga_intestazioni, column=col_valutazione)
                cell_hdr_val.value = NOME_COL_VALUTAZIONE
                cell_hdr_val.font  = Font(bold=True)

                if scrivi_lpicco_rid:
                    cell_hdr_lpicco       = ws_ag.cell(row=riga_intestazioni, column=col_lpicco_rid)
                    cell_hdr_lpicco.value = NOME_COL_LPICCO_RID
                    cell_hdr_lpicco.font  = Font(bold=True)

                # ── Righe dati: una per DPI, valori da df_dpi ────────────────
                for i in range(len(df_dpi)):
                    riga_corrente = riga_dati_base + i

                    for nome_col, offset in col_offsets.items():
                        ws_ag.cell(row=riga_corrente,
                                    column=COL_INIZIO + offset).value = \
                            df_dpi.loc[i, nome_col]

                    # Merge cella NOME_COL_MERGE
                    ws_ag.merge_cells(
                        start_row=riga_corrente, start_column=col_merge,
                        end_row=riga_corrente,   end_column=col_merge + SPAN_MERGE - 1
                    )

                    # Colore cella NOME_COL_COLORE in base al livello di rischio
                    val_colore = df_dpi.loc[i, NOME_COL_COLORE]
                    if val_colore < SOGLIA_ARANCIONE_MAX:
                        hex_colore = COLORE_ARANCIONE
                    elif val_colore > SOGLIA_ROSSO_MIN:
                        hex_colore = COLORE_ROSSO
                    elif SOGLIA_VERDE_MIN < val_colore < SOGLIA_VERDE_MAX:
                        hex_colore = COLORE_VERDE
                    else:
                        hex_colore = COLORE_GIALLO
                    ws_ag.cell(row=riga_corrente, column=col_colore).fill = \
                        PatternFill(fill_type='solid', fgColor=hex_colore)

                    val_leqa = df_dpi.loc[i, 'LeqA_rid']
                    if val_leqa < 65:
                        valutazione = 'iperprotezione'
                    elif val_leqa >= 80:
                        valutazione = 'insufficiente'
                    elif 70 <= val_leqa <= 75:
                        valutazione = 'buona'
                    else:
                        valutazione = 'accettabile'
                    ws_ag.cell(row=riga_corrente, column=col_valutazione).value = valutazione

                    # Colonna Lpicco_rid = Lpicco_max - L(DPI) - 5, colorata verde/rosso
                    if scrivi_lpicco_rid:
                        val_l_dpi   = df_dpi.loc[i, 'L']
                        lpicco_rid  = round(lpicco_max - val_l_dpi - 5, 1)
                        cell_lpicco = ws_ag.cell(row=riga_corrente, column=col_lpicco_rid)
                        cell_lpicco.value = lpicco_rid
                        if lpicco_rid < SOGLIA_LPICCO_MAX:
                            hex_lpicco = COLORE_VERDE
                        elif lpicco_rid > SOGLIA_LPICCO_MAX:
                            hex_lpicco = COLORE_ROSSO
                        else:
                            hex_lpicco = None
                        if hex_lpicco:
                            cell_lpicco.fill = PatternFill(fill_type='solid', fgColor=hex_lpicco)

                # ── Salva e chiudi ────────────────────────────────────────────
                wb_ag.save(excel_aggiornato)
                wb_ag.close()
                print(f"  → Sezione DPI scritta nel foglio '{sn}' "
                        f"(riga titolo: {riga_titolo}).")
                





        

                

        # Ri-applico i bordi per includere anche le sezioni DPI appena scritte
        exel_file.colora_bordi_celle(excel_aggiornato)

        print('\nApplicazione DPI HML completata.')

    def applica_dpi_HLM_8h(self, main_dir, dpi_dir, total_VR8h_name = "VR8h_totale"):
        '''
        Funzione che calcola il livello di esposizione giornaliero ridotto (LEX,8h) con DPI
        per ogni gruppo omogeneo, applicando il metodo HML a partire dai file prodotti da
        applica_DPI_HML(), e aggiorna il file riassuntivo VR8h_totale.

        Per ogni DPI (directory DPI_i dentro dpi_dir) e per ogni file xlsx in essa contenuto,
        calcola:
            LeqA_rid_medio = 10 * log10( sum( Ti/480 * 10^(LeqA_rid/10) ) )   [dB(A)]
            PNR_medio      = media aritmetica della colonna PNR

        e li assegna alla riga corrispondente al gruppo omogeneo nel file VR8h_totale.

        Al termine colora di verde (#008000) la cella con il valore minimo di
        LeqA_rid_medio_DPI_i per ogni riga (gruppo omogeneo), indicando il DPI
        più performante.

        INPUT:
            main_dir         = <str>, directory di lavoro principale; contiene
                               VR8h_totale.csv e riceverà VR8h_totale_dpi.xlsx
            total_VR8h_name  = <str>, nome base del file csv riassuntivo (senza estensione),
                               es. 'VR8h_totale'
            dpi_dir          = <str>, directory che contiene le sottocartelle DPI_1, DPI_2, ...
                               prodotte da applica_DPI_HML()

        OUTPUT:
            Nessun valore di ritorno. Salva VR8h_totale_dpi.xlsx in main_dir.
        '''
        import glob
        import os
        import re
        import warnings
        import numpy as np
        import openpyxl as ex
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter

        # ── Identifica le directory DPI in dpi_dir ────────────────────────────
        numero_dpi = sorted([
            d for d in os.listdir(dpi_dir)
            if os.path.isdir(os.path.join(dpi_dir, d))
        ])

        if not numero_dpi:
            print(f'Nessuna directory DPI trovata in: {dpi_dir}')
            return

        print(f'Directory DPI trovate ({len(numero_dpi)}): {numero_dpi}')

        # ── Carica VR8h_totale.csv ────────────────────────────────────────────
        tot_csv_path = os.path.join(main_dir, total_VR8h_name + '.csv')
        try:
            df_vr8h = files.read_csv(tot_csv_path)
        except Exception as e:
            print(f'Errore nel caricamento di {tot_csv_path}: {e}')
            return

        # ── Inizializza colonne PNR_medio e LeqA_rid_medio per ogni DPI ──────
        for dpi_name in numero_dpi:
            df_vr8h[f'PNR_medio_{dpi_name}']      = np.nan
            df_vr8h[f'LeqA_rid_medio_{dpi_name}'] = np.nan

        # ── Itera su ogni directory DPI ───────────────────────────────────────
        for dpi_name in numero_dpi:

            if not re.search(r'\d+', dpi_name):
                warnings.warn(f'Indice DPI non riconoscibile da: {dpi_name} — directory saltata.')
                continue

            dpi_folder = os.path.join(dpi_dir, dpi_name)
            xlsx_files = sorted(glob.glob(os.path.join(dpi_folder, '*.xlsx')))

            if not xlsx_files:
                print(f'  Nessun file xlsx trovato in: {dpi_folder}')
                continue

            print(f'\n── {dpi_name}: {len(xlsx_files)} file xlsx trovati ──')

            for xlsx_path in xlsx_files:
                fname = os.path.basename(xlsx_path)

                # Estrae nome_gruppo_omogeneo dal nome file:
                # formato atteso: VR8h_<grom>_dpi_<i>.xlsx
                m = re.search(r'VR8h_(.+)_dpi_\d+\.xlsx', fname)
                if not m:
                    warnings.warn(
                        f'Formato nome file non riconosciuto: {fname} — file saltato.'
                    )
                    continue
                grom_name = m.group(1)

                # Carica il file xlsx
                try:
                    df_xlsx = pd.read_excel(xlsx_path)
                except Exception as e:
                    warnings.warn(f'Errore nel caricamento di {fname}: {e} — file saltato.')
                    continue

                # Calcolo LeqA_rid_medio (formula LEX,8h con DPI)
                E_i = df_xlsx['Ti'] / 480 * 10 ** (df_xlsx['LeqA_rid'] / 10)
                LeqA_rid_medio = round(10 * np.log10(E_i.sum()), 1)

                # Calcolo PNR medio
                PNR_medio = round(df_xlsx['PNR'].mean(), 1)

                print(
                    f'  {fname} | GrOm={grom_name} | '
                    f'LeqA_rid_medio={LeqA_rid_medio} dB(A) | PNR_medio={PNR_medio} dB'
                )

                # Trova la riga corrispondente in VR8h_totale tramite GrOm
                mask = df_vr8h['GrOm'] == grom_name
                if not mask.any():
                    warnings.warn(
                        f'{grom_name} non trovato nel file VR8h_totale.csv'
                    )
                    continue

                df_vr8h.loc[mask, f'LeqA_rid_medio_{dpi_name}'] = LeqA_rid_medio
                df_vr8h.loc[mask, f'PNR_medio_{dpi_name}']      = PNR_medio

        # ── Salva VR8h_totale_dpi.xlsx ────────────────────────────────────────
        out_xlsx = os.path.join(main_dir, 'VR8h_totale_dpi.xlsx')
        files.write_exel(df_vr8h, out_xlsx)
        print(f'\nFile salvato: {out_xlsx}')

        # ── Colorazione: verde (#008000) sulla cella minima di ogni riga ─────
        leqa_rid_cols = [c for c in df_vr8h.columns if c.startswith('LeqA_rid_medio_')]

        if not leqa_rid_cols:
            print('Nessuna colonna LeqA_rid_medio trovata — colorazione saltata.')
            return

        wb = ex.load_workbook(out_xlsx)
        ws = wb.active

        # Mappa nome colonna -> lettera Excel
        col_map = {}
        for cell in ws[1]:
            if cell.value in leqa_rid_cols:
                col_map[cell.value] = get_column_letter(cell.column)

        green_fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')

        for row_idx in range(2, ws.max_row + 1):
            # Raccoglie i valori LeqA_rid_medio della riga corrente
            row_vals = {}
            for col_name, col_letter in col_map.items():
                val = ws[f'{col_letter}{row_idx}'].value
                if val is not None:
                    try:
                        row_vals[col_letter] = float(val)
                    except (TypeError, ValueError):
                        pass

            if not row_vals:
                continue

            # Colora di verde la cella con il valore minimo
            min_col_letter = min(row_vals, key=row_vals.get)
            ws[f'{min_col_letter}{row_idx}'].fill = green_fill

        wb.save(out_xlsx)
        print(f'Colorazione completata. Cella minima LeqA_rid_medio colorata di verde per ogni gruppo omogeneo.')
        print('\napplica_dpi_HLM_8h completata.')


      
                


                
                










        




        



         
