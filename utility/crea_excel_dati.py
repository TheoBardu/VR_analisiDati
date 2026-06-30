"""
crea_excel_dati.py
Legge i file CSV di misura del rumore e produce un file Excel
con la struttura definita in esempio_export.xlsx.

Uso:
    python crea_excel_dati.py <data_dir> <main_dir> [--output nome_file.xlsx]

Argomenti:
    data_dir  Directory contenente averaged_data.csv e mis*.csv
    main_dir  Directory con scheda_gruppi_dpi.xlsx; qui viene salvato l'output
"""

import argparse
import os

import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string

SCHEDA_MANSIONI = "Scheda_mansioni"
OUTPUT_DEFAULT = "dati_misure.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _duration_to_seconds(duration_str):
    """Converte 'HH:MM:SS' in secondi interi. Ritorna None se non parsabile."""
    try:
        parts = str(duration_str).split(":")
        h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
        return h * 3600 + m * 60 + s
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Lettura dati
# ---------------------------------------------------------------------------

def load_averaged_data(data_dir):
    path = os.path.join(data_dir, "averaged_data.csv")
    if not os.path.exists(path):
        raise FileNotFoundError(f"File non trovato: {path}")
    df = pd.read_csv(path, index_col=0)
    return df


def load_mis_files(data_dir):
    """Carica e concatena tutti i file mis*.csv presenti in data_dir."""
    dfs = []
    for fname in sorted(os.listdir(data_dir)):
        if fname.startswith("mis") and fname.endswith(".csv"):
            fpath = os.path.join(data_dir, fname)
            df = pd.read_csv(fpath, index_col=0)
            dfs.append(df)
    if not dfs:
        return pd.DataFrame()
    return pd.concat(dfs, ignore_index=True)


def load_scheda(main_dir):
    path = os.path.join(main_dir, "scheda_gruppi_dpi.xlsx")
    if not os.path.exists(path):
        raise FileNotFoundError(f"File non trovato: {path}")
    df = pd.read_excel(path, sheet_name=SCHEDA_MANSIONI, header=1)
    return df


# ---------------------------------------------------------------------------
# Scrittura Excel
# ---------------------------------------------------------------------------

def _write_headers(ws):
    """Scrive le due righe di intestazione con le celle unite."""

    # --- Row 1 merged cells ---
    ws.merge_cells("A1:A2")
    ws["A1"] = "ID"

    ws.merge_cells("B1:B2")
    ws["B1"] = "Descrizione"

    # C1 lasciata vuota, C2 = 'U1a,m' (non merged)
    ws["C2"] = "U1a,m"

    ws.merge_cells("D1:D2")
    ws["D1"] = "PPEAK(C)\n+\nU(Lpicco,C)"

    ws.merge_cells("E1:E2")
    ws["E1"] = "LAEQ,Tp"

    ws.merge_cells("F1:F2")
    ws["F1"] = "LCEQ,TP"

    # MISURA n°1-6: tutte con 4 colonne (Sec, LAeq, LCeq, Lpicco)
    track_ranges = [
        ("G", "J"), ("K", "N"), ("O", "R"),
        ("S", "V"), ("W", "Z"), ("AA", "AD"),
    ]
    for i, (start_col, end_col) in enumerate(track_ranges, start=1):
        ws.merge_cells(f"{start_col}1:{end_col}1")
        ws[f"{start_col}1"] = f"MISURA n°{i}"
        c = column_index_from_string(start_col)
        ws.cell(row=2, column=c,     value="Sec.")
        ws.cell(row=2, column=c + 1, value="LAeq,T ")
        ws.cell(row=2, column=c + 2, value="LCeq,T ")
        ws.cell(row=2, column=c + 3, value="Lpicco,C")

    # Row 3: separatore
    ws["A3"] = "⌂"


def _get_descrizione(scheda_df, id_misura):
    rows = scheda_df[scheda_df["ID_misura"] == id_misura]
    if not rows.empty:
        return rows.iloc[0].get("Descrizione_compito", "")
    return ""


def _get_track(mis_df, letter_id, n_track):
    """Ritorna dict con i valori di una traccia, o None se non trovata."""
    if mis_df.empty:
        return None
    mask = (mis_df["letter_ID"] == letter_id) & (mis_df["nTrack"] == n_track)
    rows = mis_df[mask]
    if rows.empty:
        return None
    r = rows.iloc[0]
    return {
        "sec":   _duration_to_seconds(r.get("durata", "")),
        "leqA":  r.get("LeqA_eq",   None),
        "leqC":  r.get("LeqC_eq",   None),
        "ppeak": r.get("PeakC_max", None),
    }


def write_excel(df_avg, df_mis, df_scheda, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dati"

    _write_headers(ws)

    # Colonne di partenza per le 6 tracce (indice 1-based), 4 colonne ciascuna
    # Track 1: G=7  Track 2: K=11  Track 3: O=15
    # Track 4: S=19 Track 5: W=23  Track 6: AA=27
    track_start_cols = [7, 11, 15, 19, 23, 27]

    excel_row = 4  # i dati partono da riga 4
    for _, avg_row in df_avg.iterrows():
        id_misura = avg_row["ID"]

        ws.cell(row=excel_row, column=1,  value=id_misura)
        ws.cell(row=excel_row, column=2,  value=_get_descrizione(df_scheda, id_misura))
        ws.cell(row=excel_row, column=3,  value=avg_row.get("U",     None))
        ws.cell(row=excel_row, column=4,  value=avg_row.get("Ppeak", None))
        ws.cell(row=excel_row, column=5,  value=avg_row.get("LeqA",  None))
        ws.cell(row=excel_row, column=6,  value=avg_row.get("LeqC",  None))

        # Tutte le tracce: 4 colonne ciascuna (Sec, LAeq, LCeq, Lpicco)
        for track_n, col_start in zip(range(1, 7), track_start_cols):
            t = _get_track(df_mis, id_misura, track_n)
            if t:
                ws.cell(row=excel_row, column=col_start,     value=t["sec"])
                ws.cell(row=excel_row, column=col_start + 1, value=t["leqA"])
                ws.cell(row=excel_row, column=col_start + 2, value=t["leqC"])
                ws.cell(row=excel_row, column=col_start + 3, value=t["ppeak"])

        excel_row += 1

    wb.save(output_path)
    print(f"File salvato: {output_path}  ({excel_row - 4} righe dati)")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Crea file Excel con dati di misura del rumore"
    )
    parser.add_argument(
        "data_dir",
        help="Directory contenente averaged_data.csv e mis*.csv",
    )
    parser.add_argument(
        "main_dir",
        help="Directory con scheda_gruppi_dpi.xlsx; output salvato qui",
    )
    parser.add_argument(
        "--output",
        default=OUTPUT_DEFAULT,
        help=f"Nome file di output (default: {OUTPUT_DEFAULT})",
    )
    args = parser.parse_args()

    print("Lettura dati...")
    df_avg    = load_averaged_data(args.data_dir)
    df_mis    = load_mis_files(args.data_dir)
    df_scheda = load_scheda(args.main_dir)

    print(f"  averaged_data: {len(df_avg)} misure")
    print(f"  mis files:     {len(df_mis)} tracce totali")
    print(f"  scheda:        {len(df_scheda)} righe")

    output_path = os.path.join(args.main_dir, args.output)
    write_excel(df_avg, df_mis, df_scheda, output_path)


if __name__ == "__main__":
    main()
