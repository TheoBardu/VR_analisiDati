import math
from copy import copy
from datetime import datetime, timedelta

import openpyxl

file2extend = '/Users/theo/Desktop/P.IVA/Aziende/Ermes/Lavori/DEFRANCESCHI/rev/rev0/Rumore/misure/misH/730 A10574-26070315-110407.xlsx'
Ttot = 6  # durata minima desiderata in minuti

SHEET_NAME = "Profilo storico"
TIME_FMT = "%Y-%m-%d  %H:%M:%S"
TIME_COL_HEADER = "Data/Tempo"
TIPO_COL_HEADER = "Tipo di registrazione"
STOP_MARKER = "Fermare"

wb = openpyxl.load_workbook(file2extend)
ws = wb[SHEET_NAME]

header_row = 1
max_col = ws.max_column


def find_col(header_substring):
    for col in range(1, max_col + 1):
        header_val = ws.cell(row=header_row, column=col).value
        if header_val and header_substring in header_val:
            return col
    return None


time_col = find_col(TIME_COL_HEADER)
tipo_col = find_col(TIPO_COL_HEADER)

if time_col is None:
    raise ValueError(f"Colonna '{TIME_COL_HEADER}' non trovata nel foglio '{SHEET_NAME}'.")

# Trova l'ultima riga dati contigua a partire dalla riga successiva all'intestazione
data_start = header_row + 1
data_end = data_start
while ws.cell(row=data_end, column=time_col).value is not None:
    data_end += 1
data_end -= 1

if data_end < data_start:
    raise ValueError(f"Nessuna riga dati trovata nel foglio '{SHEET_NAME}'.")

# La riga di stop ("Fermare") non ha misure: va tolta dai blocchi duplicati e
# riscritta una sola volta, alla fine, per non lasciare righe vuote tra i blocchi.
has_stop_row = tipo_col is not None and ws.cell(row=data_end, column=tipo_col).value == STOP_MARKER
stop_row = data_end if has_stop_row else None
active_end = data_end - 1 if has_stop_row else data_end

data_rows = list(range(data_start, active_end + 1))
footer_rows = list(range(data_end + 1, ws.max_row + 1))

first_time = datetime.strptime(ws.cell(row=data_start, column=time_col).value, TIME_FMT)
last_time = datetime.strptime(ws.cell(row=data_end, column=time_col).value, TIME_FMT)

deltaT = (last_time - first_time).total_seconds()   # usato per calcolare N
# Se c'è una riga di stop, la sua timestamp coincide col punto in cui deve
# iniziare il blocco successivo: nessuno shift aggiuntivo, continuità totale.
deltaT_shift = deltaT if has_stop_row else deltaT + 1

N = math.ceil(Ttot * 60 / deltaT)

def snapshot_row(r):
    """Salva valore e stile di ogni cella della riga, per poterla riscrivere altrove."""
    snap = [
        (
            cell.value, copy(cell.font), copy(cell.fill), copy(cell.border),
            copy(cell.alignment), copy(cell.protection), cell.number_format,
        )
        for cell in (ws.cell(row=r, column=col) for col in range(1, max_col + 1))
    ]
    return snap, ws.row_dimensions[r].height


def write_row(row_num, snap, height):
    for col, (value, font, fill, border, alignment, protection, number_format) in enumerate(snap, start=1):
        cell = ws.cell(row=row_num, column=col)
        cell.value = value
        cell.font = font
        cell.fill = fill
        cell.border = border
        cell.alignment = alignment
        cell.protection = protection
        cell.number_format = number_format
    ws.row_dimensions[row_num].height = height


# Salva valori/stili della riga di stop (se presente) e delle righe di footer
# prima di cancellarle dal foglio: verranno riscritte dopo i blocchi aggiunti.
stop_snapshot = snapshot_row(stop_row) if has_stop_row else None
footer_snapshot = [snapshot_row(r) for r in footer_rows]

delete_from = stop_row if has_stop_row else (footer_rows[0] if footer_rows else None)
if delete_from is not None:
    ws.delete_rows(delete_from, ws.max_row - delete_from + 1)

# Genera le righe estese modificando solo la colonna Data/Tempo
next_row = active_end + 1
for i in range(1, N + 1):
    for j, src_row in enumerate(data_rows):
        new_time = first_time + timedelta(seconds=i * deltaT_shift + j)
        for col in range(1, max_col + 1):
            src_cell = ws.cell(row=src_row, column=col)
            new_cell = ws.cell(row=next_row, column=col)
            new_cell.value = new_time.strftime(TIME_FMT) if col == time_col else src_cell.value
            new_cell.font = copy(src_cell.font)
            new_cell.fill = copy(src_cell.fill)
            new_cell.border = copy(src_cell.border)
            new_cell.alignment = copy(src_cell.alignment)
            new_cell.protection = copy(src_cell.protection)
            new_cell.number_format = src_cell.number_format
        ws.row_dimensions[next_row].height = ws.row_dimensions[src_row].height
        next_row += 1

# Riscrive la riga di stop originale una sola volta, con il tempo aggiornato,
# a marcare la vera fine della traccia estesa (nessuna riga vuota tra i blocchi)
if has_stop_row:
    snap, height = stop_snapshot
    snap = list(snap)
    stop_time = first_time + timedelta(seconds=N * deltaT_shift + deltaT)
    value, font, fill, border, alignment, protection, number_format = snap[time_col - 1]
    snap[time_col - 1] = (stop_time.strftime(TIME_FMT), font, fill, border, alignment, protection, number_format)
    write_row(next_row, snap, height)
    next_row += 1

# Riscrive le righe di footer originali dopo i nuovi blocchi
for snap, height in footer_snapshot:
    write_row(next_row, snap, height)
    next_row += 1

# Sovrascrive il file originale
wb.save(file2extend)

total_seconds = N * deltaT_shift + deltaT
print(f"File sovrascritto: {file2extend}")
print(f"deltaT = {int(deltaT)}s | copie aggiunte = {N} | durata totale ≈ {total_seconds/60:.1f} min")
