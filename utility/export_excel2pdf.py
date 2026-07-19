"""
export_excel2pdf.py
Apre il file excel indicato (tipicamente VR8h_totale_aggiornato.xlsx prodotto da
analisi_8h()), imposta l'area di stampa di ogni foglio sulle celle popolate ed
esporta TUTTI i fogli in un UNICO file PDF.

Prima dell'export le colonne vengono allargate per adattarsi al contenuto, cosi'
il testo non risulta troncato nel PDF.

L'export usa LibreOffice (soffice) in modalita' headless: nessuna finestra viene
aperta e tutti i fogli finiscono in un unico PDF rispettando le aree di stampa.

NOTA: Microsoft Excel per Mac (16.x) non e' utilizzabile come motore alternativo
perche' non espone piu' alcun comando AppleScript per l'export in PDF
('do Visual Basic' e' stato rimosso dal dizionario, 'save workbook as' rifiuta il
formato PDF e 'print out' non accetta un percorso di destinazione).

Uso:
    python export_excel2pdf.py <excel_document> [--output file.pdf]

Argomenti:
    excel_document  Percorso del file .xlsx da esportare
"""

import argparse
import os
import shutil
import subprocess
import tempfile

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

# ---------------------------------------------------------------------------
# Configurazione
# ---------------------------------------------------------------------------

# Impaginazione applicata a ogni foglio prima dell'export
ORIENTAMENTO = 'landscape'   # 'landscape' oppure 'portrait'
MARGINI = 0.3                # pollici, margini laterali/superiori
TIMEOUT_EXPORT = 180         # secondi

# Adattamento larghezza colonne al contenuto
LARGHEZZA_MIN = 4            # caratteri
LARGHEZZA_MAX = 40           # caratteri, evita colonne spropositate
LARGHEZZA_PADDING = 1.2      # caratteri di margine aggiunti al contenuto


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _ultima_riga_popolata(ws):
    """Ultima riga del foglio con almeno una cella valorizzata."""
    last = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                last = cell.row
                break
    return last


def _ultima_col_popolata(ws):
    """Ultima colonna del foglio con almeno una cella valorizzata."""
    last = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.column > last:
                last = cell.column
    return last


def _foglio_vuoto(ws):
    """True se il foglio non contiene alcun valore."""
    return ws.max_row == 1 and ws.max_column == 1 and ws['A1'].value is None


def _adatta_larghezze(ws):
    """
    Allarga le colonne in base al contenuto piu' lungo, cosi' il testo non
    risulta troncato nel PDF. Le celle unite vengono ignorate: il loro testo si
    distribuisce su piu' colonne e falserebbe il calcolo.
    """
    celle_unite = set()
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                celle_unite.add((r, c))

    larghezze = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or (cell.row, cell.column) in celle_unite:
                continue
            lung = len(str(cell.value))
            if lung > larghezze.get(cell.column, 0):
                larghezze[cell.column] = lung

    for col, lung in larghezze.items():
        letter = get_column_letter(col)
        larghezza = min(max(lung + LARGHEZZA_PADDING, LARGHEZZA_MIN), LARGHEZZA_MAX)
        attuale = ws.column_dimensions[letter].width
        # Allargo soltanto: non restringo colonne gia' impostate a mano
        if attuale is None or attuale < larghezza:
            ws.column_dimensions[letter].width = larghezza


# ---------------------------------------------------------------------------
# Area di stampa
# ---------------------------------------------------------------------------

def imposta_aree_di_stampa(excel_document):
    """
    Imposta su ogni foglio l'area di stampa pari al rettangolo di celle popolate
    (A1 -> ultima riga/colonna con contenuto), adatta le larghezze delle colonne
    al contenuto e imposta l'impaginazione: A4, orizzontale, adattato alla
    larghezza della pagina.

    Sui fogli 'Scheda N' la riga di intestazione viene ripetuta su ogni pagina.
    Il file viene salvato in place.

    Parametri
    ----------
    excel_document : str
        Percorso del file .xlsx da preparare

    Ritorna
    -------
    dict : {nome_foglio: area_di_stampa} per i fogli effettivamente impostati
    """
    wb = openpyxl.load_workbook(excel_document)
    aree = {}

    for ws in wb.worksheets:
        if _foglio_vuoto(ws):
            print(f"  '{ws.title}': foglio vuoto, saltato")
            continue

        max_row = _ultima_riga_popolata(ws)
        max_col = _ultima_col_popolata(ws)
        if max_row == 0 or max_col == 0:
            continue

        area = f"A1:{get_column_letter(max_col)}{max_row}"
        ws.print_area = area
        aree[ws.title] = area

        _adatta_larghezze(ws)

        # Impaginazione: A4 orizzontale, adattato alla larghezza
        ws.page_setup.orientation = ORIENTAMENTO
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        # Senza pageSetUpPr openpyxl ignora i parametri fitToWidth/fitToHeight
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        ws.page_margins.left = MARGINI
        ws.page_margins.right = MARGINI
        ws.page_margins.top = MARGINI
        ws.page_margins.bottom = MARGINI

        # Ripeti l'intestazione della tabella su ogni pagina
        if ws.title.strip().lower().startswith('scheda'):
            ws.print_title_rows = '1:1'

        print(f"  '{ws.title}': area di stampa {area}")

    wb.save(excel_document)
    return aree


# ---------------------------------------------------------------------------
# Motori di export
# ---------------------------------------------------------------------------

def _export_libreoffice(excel_document, pdf_output):
    """Esporta in PDF con LibreOffice headless (nessuna finestra viene aperta)."""
    soffice = shutil.which('soffice') or shutil.which('libreoffice')
    if soffice is None:
        raise RuntimeError(
            "LibreOffice non trovato nel PATH (cercati 'soffice' e 'libreoffice').\n"
            "Installalo oppure usa --tipologia excel."
        )

    # soffice nomina il PDF come il sorgente: converto in una dir temporanea
    # e poi sposto il risultato sul percorso richiesto.
    with tempfile.TemporaryDirectory() as tmpdir:
        cmd = [
            soffice,
            '--headless',
            '--norestore',
            '--convert-to', 'pdf:calc_pdf_Export',
            '--outdir', tmpdir,
            excel_document,
        ]
        res = subprocess.run(cmd, capture_output=True, text=True, timeout=TIMEOUT_EXPORT)

        if res.returncode != 0:
            raise RuntimeError(
                f"LibreOffice ha restituito un errore (codice {res.returncode}):\n"
                f"{res.stderr.strip() or res.stdout.strip()}"
            )

        nome_pdf = os.path.splitext(os.path.basename(excel_document))[0] + '.pdf'
        prodotto = os.path.join(tmpdir, nome_pdf)
        if not os.path.exists(prodotto):
            raise RuntimeError(
                f"LibreOffice non ha prodotto il PDF atteso ({nome_pdf}).\n"
                f"{res.stdout.strip()}"
            )

        shutil.move(prodotto, pdf_output)


# ---------------------------------------------------------------------------
# Funzione principale
# ---------------------------------------------------------------------------

def esporta_pdf(excel_document, pdf_output=None):
    """
    Imposta le aree di stampa ed esporta tutti i fogli dell'excel in un unico PDF.

    Parametri
    ----------
    excel_document : str
        Percorso del file .xlsx da esportare
    pdf_output : str, opzionale
        Percorso del PDF di destinazione. Se None viene usato lo stesso
        percorso/nome del sorgente con estensione .pdf

    Ritorna
    -------
    str : percorso del PDF prodotto
    """
    excel_document = os.path.abspath(os.path.expanduser(excel_document))

    if not os.path.exists(excel_document):
        raise FileNotFoundError(f"File excel non trovato: {excel_document}")

    if pdf_output is None:
        pdf_output = os.path.splitext(excel_document)[0] + '.pdf'
    pdf_output = os.path.abspath(os.path.expanduser(pdf_output))

    print(f"Impostazione aree di stampa in {os.path.basename(excel_document)}...")
    imposta_aree_di_stampa(excel_document)

    print("Export PDF con LibreOffice...")
    _export_libreoffice(excel_document, pdf_output)

    print(f"PDF creato: {pdf_output}")
    return pdf_output


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Esporta in un unico PDF tutti i fogli di un file excel, "
                    "rispettando l'area di stampa"
    )
    parser.add_argument(
        "excel_document",
        help="Percorso del file .xlsx da esportare",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Percorso del PDF di output (default: stesso nome del file excel)",
    )
    args = parser.parse_args()

    esporta_pdf(args.excel_document, args.output)


if __name__ == "__main__":
    main()
